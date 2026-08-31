from django.shortcuts import render, redirect, get_object_or_404
import csv
import json
import re
import unicodedata
from datetime import timedelta, datetime, date, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from .models import Cuadrilla
from .models import SoporteTecnico
from .models import AsignacionMaterial
from .models import BajaMaterial
from .models import EquipoRetiradoManual
from .models import AsignacionIdentificador
from .models import BajaIdentificador
from .models import Cliente
from .forms import AsignacionMaterialForm
from .forms import (
    CuadrillaForm,
    AsignacionMaterialForm,
    CrearCuadrillaConResponsableForm,
    SoporteTecnicoForm,
    CrearSoporteConResponsableForm,
)
from inventario.models import Producto, DetalleIdentificador, Proveedor
from inventario.models import Categoria, Marca, ModeloEquipo
from django.contrib import messages
from collections import defaultdict
from usuarios.decorators import role_required
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from inventario.forms import es_categoria_activa, es_categoria_fibra, es_categoria_codigo_unico


def _parse_decimal_input(valor):
    texto = str(valor or '').strip().replace(',', '.')
    if not texto:
        return Decimal('0')
    try:
        numero = Decimal(texto)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')
    return numero if numero >= 0 else Decimal('0')


def _producto_es_fibra(producto):
    categoria_nombre = producto.categoria.nombre if producto.categoria else ''
    return (
        es_categoria_fibra(categoria_nombre)
        or bool(getattr(producto, 'tipo_fibra_id', None))
        or bool(getattr(producto, 'hilo', None))
    )


def _producto_es_adss(producto):
    nombre_producto = (getattr(producto, 'nombre', '') or '').lower()
    codigo_producto = (getattr(producto, 'codigo', '') or '').lower()
    categoria_nombre = (producto.categoria.nombre if producto.categoria else '').lower()
    tipo_fibra_nombre = (getattr(getattr(producto, 'tipo_fibra', None), 'nombre', '') or '').lower()
    return 'adss' in nombre_producto or 'adss' in codigo_producto or 'adss' in categoria_nombre or 'adss' in tipo_fibra_nombre


def _usuario_es_cuadrilla(usuario):
    return usuario.groups.filter(name__icontains='cuadrilla').exists() or getattr(usuario, 'is_cuadrilla', False)


def _usuario_es_soporte(usuario):
    return usuario.groups.filter(name__icontains='soporte').exists()


def _filtro_cuadrilla_por_usuario(usuario):
    nombres_cuadrilla = set()
    for grupo in usuario.groups.filter(name__icontains='cuadrilla'):
        nombre_grupo = (grupo.name or '').strip()
        nombre_sin_prefijo = nombre_grupo.lower().replace('cuadrilla', '').strip()

        if nombre_grupo:
            nombres_cuadrilla.add(nombre_grupo)
        if nombre_sin_prefijo:
            nombres_cuadrilla.add(nombre_sin_prefijo.upper())
            nombres_cuadrilla.add(nombre_sin_prefijo)

    if usuario.username:
        nombres_cuadrilla.add(usuario.username)

    filtro = Q(cuadrilla__responsable=usuario)
    for nombre in nombres_cuadrilla:
        filtro |= Q(cuadrilla__nombre__iexact=nombre)
    return filtro


def _normalizar_cliente_nombre(nombre):
    # Colapsa espacios repetidos y estandariza formato para evitar duplicados.
    texto = ' '.join(str(nombre or '').strip().split())
    return texto.title()


def _normalizar_texto(valor):
    return ' '.join(str(valor or '').strip().split())


def _extraer_campos_observacion(observacion):
    texto = str(observacion or '').strip()
    resultado = {
        'coordenadas': '',
        'sector': '',
        'coord_punta_inicial': '',
        'coord_punta_final': '',
        'coord_cliente': '',
        'coord_caja': '',
    }

    if not texto:
        return resultado

    for parte in [p.strip() for p in texto.split('|') if p.strip()]:
        if ':' not in parte:
            continue

        clave_raw, valor_raw = parte.split(':', 1)
        clave = _normalizar_texto(clave_raw).lower()
        valor = _normalizar_texto(valor_raw)

        if clave == 'coordenadas':
            resultado['coordenadas'] = valor
        elif clave == 'sector':
            resultado['sector'] = valor
        elif clave in ['coord punta inicial', 'coordenada punta inicial', 'coordenadas punta inicial']:
            resultado['coord_punta_inicial'] = valor
        elif clave in ['coord punta final', 'coordenada punta final', 'coordenadas punta final']:
            resultado['coord_punta_final'] = valor
        elif clave in ['coord cliente', 'coordenada cliente', 'coordenadas cliente']:
            resultado['coord_cliente'] = valor
        elif clave in ['coord caja', 'coordenada caja', 'coordenadas caja']:
            resultado['coord_caja'] = valor

    return resultado


def _detalle_observacion_sin_coordenadas(observacion):
    texto = str(observacion or '').strip()
    if not texto:
        return ''

    claves_coord = {
        'coordenadas',
        'sector',
        'coord punta inicial',
        'coordenada punta inicial',
        'coordenadas punta inicial',
        'coord punta final',
        'coordenada punta final',
        'coordenadas punta final',
        'coord cliente',
        'coordenada cliente',
        'coordenadas cliente',
        'coord caja',
        'coordenada caja',
        'coordenadas caja',
    }

    partes_visibles = []
    for parte in [p.strip() for p in texto.split('|') if p.strip()]:
        if ':' not in parte:
            partes_visibles.append(parte)
            continue

        clave_raw, _ = parte.split(':', 1)
        clave = _normalizar_texto(clave_raw).lower()
        if clave in claves_coord:
            continue
        partes_visibles.append(parte)

    return ' | '.join(partes_visibles)


def _extraer_coordenadas_sector(observacion):
    campos = _extraer_campos_observacion(observacion)
    return campos.get('coordenadas', ''), campos.get('sector', '')


def _filtrar_historial_construccion_qs(request, bajas_qs):
    sector_q = _normalizar_texto(request.GET.get('sector'))
    producto_q = _normalizar_texto(request.GET.get('producto'))
    arreglo_q = (request.GET.get('arreglo') or '').strip()
    fecha_q = (request.GET.get('fecha') or '').strip()

    if producto_q:
        bajas_qs = bajas_qs.filter(
            Q(asignacion__producto__nombre__icontains=producto_q) |
            Q(asignacion__producto__codigo__icontains=producto_q) |
            Q(producto_nombre__icontains=producto_q) |
            Q(producto_codigo__icontains=producto_q)
        )

    if arreglo_q:
        bajas_qs = bajas_qs.filter(tipo_uso=arreglo_q)

    if fecha_q:
        try:
            fecha_obj = datetime.strptime(fecha_q, '%Y-%m-%d').date()
            bajas_qs = bajas_qs.filter(fecha__date=fecha_obj)
        except ValueError:
            pass

    if sector_q:
        bajas_qs = bajas_qs.filter(observacion__icontains=sector_q)

    return bajas_qs, {
        'sector': sector_q,
        'producto': producto_q,
        'arreglo': arreglo_q,
        'fecha': fecha_q,
    }


def _normalizar_header(texto):
    base = _normalizar_texto(texto).lower()
    base = ''.join(
        ch for ch in unicodedata.normalize('NFD', base)
        if unicodedata.category(ch) != 'Mn'
    )
    return base.replace('_', ' ').replace('-', ' ').strip()


def _mapear_columnas(header_row):
    header_map = {}
    for idx, val in enumerate(header_row):
        normalizado = _normalizar_header(val)
        if not normalizado:
            continue
        header_map[normalizado] = idx

    idx_nombre = None
    for clave in ('nombre', 'nombre cliente', 'cliente', 'clientes'):
        if clave in header_map:
            idx_nombre = header_map[clave]
            break

    if idx_nombre is None:
        return None

    idx_ct = None
    for clave in ('ct', 'codigo ct', 'cod ct'):
        if clave in header_map:
            idx_ct = header_map[clave]
            break

    idx_cedula = None
    for clave in ('cedula', 'cedula cliente', 'dni', 'documento', 'numero cedula', 'num cedula'):
        if clave in header_map:
            idx_cedula = header_map[clave]
            break

    return {
        'nombre': idx_nombre,
        'ct': idx_ct,
        'cedula': idx_cedula,
    }


def _extraer_cliente_por_indices(row, columnas):
    idx_nombre = columnas.get('nombre')
    idx_ct = columnas.get('ct')
    idx_cedula = columnas.get('cedula')

    nombre_raw = row[idx_nombre] if idx_nombre is not None and idx_nombre < len(row) else ''
    ct_raw = row[idx_ct] if idx_ct is not None and idx_ct < len(row) else ''
    cedula_raw = row[idx_cedula] if idx_cedula is not None and idx_cedula < len(row) else ''

    nombre = _normalizar_cliente_nombre(nombre_raw)
    return {
        'nombre': nombre,
        'ct': _normalizar_texto(ct_raw),
        'cedula': _normalizar_texto(cedula_raw),
    }


def _es_mac(valor):
    texto = (valor or '').strip().upper()
    return bool(re.fullmatch(r'[0-9A-F]{2}([-:][0-9A-F]{2}){5}', texto))


def _inferir_tipo_identificador(valor):
    return 'MAC' if _es_mac(valor) else 'SERIAL'


def _leer_clientes_archivo(archivo):
    extension = Path(archivo.name).suffix.lower()
    clientes = []

    if extension == '.csv':
        contenido = archivo.read().decode('utf-8-sig', errors='ignore').splitlines()
        rows = list(csv.reader(contenido))
        if not rows:
            return clientes

        columnas = _mapear_columnas(rows[0])
        data_rows = rows[1:] if columnas else rows
        if not columnas:
            columnas = {'nombre': 0, 'ct': None, 'cedula': None}

        for row in data_rows:
            if not row:
                continue
            cliente = _extraer_cliente_por_indices(row, columnas)
            if cliente['nombre']:
                clientes.append(cliente)
        return clientes

    if extension == '.xlsx':
        from openpyxl import load_workbook

        workbook = load_workbook(archivo, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True) if row]
        if not rows:
            return clientes

        columnas = _mapear_columnas(rows[0])
        data_rows = rows[1:] if columnas else rows
        if not columnas:
            columnas = {'nombre': 0, 'ct': None, 'cedula': None}

        for row in data_rows:
            cliente = _extraer_cliente_por_indices(row, columnas)
            if cliente['nombre']:
                clientes.append(cliente)
        return clientes

    raise ValueError('Formato no soportado. Use .csv o .xlsx')


def _parse_fecha_retiro(valor):
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime.combine(valor, datetime.min.time())

    texto = _normalizar_texto(valor)
    if not texto:
        return timezone.now()

    for formato in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            continue

    return timezone.now()


def _buscar_producto_retiro(codigo, nombre):
    codigo = _normalizar_texto(codigo)
    nombre = _normalizar_texto(nombre)

    if codigo:
        producto = Producto.objects.filter(codigo__iexact=codigo).first()
        if producto:
            return producto

    if nombre:
        producto = Producto.objects.filter(nombre__iexact=nombre).first()
        if producto:
            return producto

    return None


def _normalizar_header_retiro(texto):
    base = _normalizar_header(texto)
    return base.replace('  ', ' ')


def _mapear_columnas_retiro(header_row):
    header_map = {}
    for idx, val in enumerate(header_row):
        normalizado = _normalizar_header_retiro(val)
        if normalizado:
            header_map[normalizado] = idx

    def buscar(*claves):
        for clave in claves:
            if clave in header_map:
                return header_map[clave]
        return None

    return {
        'fecha': buscar('fecha', 'fecha retiro', 'fecha de retiro'),
        'producto': buscar('producto', 'nombre producto', 'producto nuevo'),
        'codigo': buscar('codigo', 'código', 'codigo producto', 'codigo nuevo', 'codigo equipo'),
        'proveedor': buscar('proveedor', 'nombre proveedor'),
        'cuadrilla': buscar('cuadrilla', 'nombre cuadrilla'),
        'cliente': buscar('cliente', 'nombre cliente'),
        'ct': buscar('ct', 'codigo ct', 'cod ct'),
        'detalle': buscar('detalle', 'detalle cambio', 'observacion', 'observación'),
        'serial_mac': buscar('serial', 'mac', 'serial/mac', 'serial o mac', 'serial mac', 'serial_mac'),
        'codigo_retirado': buscar('codigo retirado', 'codigo_equipo', 'codigo equipo', 'codigo unico'),
        'categoria': buscar('categoria', 'categoria retirado'),
        'marca': buscar('marca', 'marca retirado'),
        'modelo': buscar('modelo', 'modelo retirado'),
        'cantidad': buscar('cantidad'),
        'metraje': buscar('metraje'),
        'origen': buscar('origen', 'tipo', 'tipo origen'),
    }


def _leer_retiros_manual_archivo(archivo):
    extension = Path(archivo.name).suffix.lower()
    registros = []

    def extraer_row(row, columnas):
        def val(clave):
            idx = columnas.get(clave)
            return row[idx] if idx is not None and idx < len(row) else ''

        fecha = _parse_fecha_retiro(val('fecha'))
        producto_nombre = _normalizar_texto(val('producto'))
        producto_codigo = _normalizar_texto(val('codigo'))
        proveedor_nombre = _normalizar_texto(val('proveedor'))
        cuadrilla_nombre = _normalizar_texto(val('cuadrilla'))
        cliente_nombre = _normalizar_cliente_nombre(val('cliente'))
        ct_cliente = _normalizar_texto(val('ct'))
        detalle_cambio = _normalizar_texto(val('detalle'))
        serial_mac_retirado = _normalizar_texto(val('serial_mac')).upper()
        codigo_retirado = _normalizar_texto(val('codigo_retirado')).upper()
        categoria_retirado = _normalizar_texto(val('categoria'))
        marca_retirado = _normalizar_texto(val('marca'))
        modelo_retirado = _normalizar_texto(val('modelo'))
        origen = _normalizar_texto(val('origen')).title() or 'Manual'
        if origen not in {'Cambio', 'Manual'}:
            origen = 'Manual'
        try:
            cantidad = int(_normalizar_texto(val('cantidad')) or '1')
        except ValueError:
            cantidad = 1
        metraje = _parse_decimal_input(val('metraje'))

        producto = _buscar_producto_retiro(producto_codigo, producto_nombre)
        registros.append({
            'fecha': fecha,
            'origen': origen,
            'producto': producto,
            'producto_nombre': producto_nombre or (producto.nombre if producto else ''),
            'producto_codigo': producto_codigo or (producto.codigo if producto else ''),
            'proveedor_nombre': proveedor_nombre or (producto.proveedor.nombre if producto and producto.proveedor else ''),
            'cuadrilla_nombre': cuadrilla_nombre,
            'cliente_nombre': cliente_nombre,
            'ct_cliente': ct_cliente,
            'detalle_cambio': detalle_cambio,
            'serial_mac_retirado': serial_mac_retirado,
            'codigo_retirado': codigo_retirado,
            'categoria_retirado': categoria_retirado,
            'marca_retirado': marca_retirado,
            'modelo_retirado': modelo_retirado,
            'cantidad': cantidad if cantidad > 0 else 1,
            'metraje': metraje,
        })

    if extension == '.csv':
        contenido = archivo.read().decode('utf-8-sig', errors='ignore').splitlines()
        rows = list(csv.reader(contenido))
        if not rows:
            return registros
        columnas = _mapear_columnas_retiro(rows[0])
        data_rows = rows[1:] if any(v is not None for v in columnas.values()) else rows
        if not any(v is not None for v in columnas.values()):
            columnas = {
                'fecha': 0,
                'producto': 1,
                'codigo': 2,
                'proveedor': 3,
                'cuadrilla': 4,
                'cliente': 5,
                'ct': 6,
                'detalle': 7,
                'serial_mac': 8,
                'codigo_retirado': 9,
                'categoria': 10,
                'marca': 11,
                'modelo': 12,
                'cantidad': 13,
                'metraje': 14,
                'origen': 15,
            }
        for row in data_rows:
            if any(str(c or '').strip() for c in row):
                extraer_row(row, columnas)
        return registros

    if extension == '.xlsx':
        from openpyxl import load_workbook

        workbook = load_workbook(archivo, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True) if row]
        if not rows:
            return registros
        columnas = _mapear_columnas_retiro(rows[0])
        data_rows = rows[1:] if any(v is not None for v in columnas.values()) else rows
        if not any(v is not None for v in columnas.values()):
            columnas = {
                'fecha': 0,
                'producto': 1,
                'codigo': 2,
                'proveedor': 3,
                'cuadrilla': 4,
                'cliente': 5,
                'ct': 6,
                'detalle': 7,
                'serial_mac': 8,
                'codigo_retirado': 9,
                'categoria': 10,
                'marca': 11,
                'modelo': 12,
                'cantidad': 13,
                'metraje': 14,
                'origen': 15,
            }
        for row in data_rows:
            if any(str(c or '').strip() for c in row):
                extraer_row(row, columnas)
        return registros

    raise ValueError('Formato no soportado. Use .csv o .xlsx')

@role_required('admin', 'bodega')
def lista_cuadrillas(request):

    cuadrillas = Cuadrilla.objects.all()

    return render(request, 'cuadrillas/cuadrillas.html', {
        'cuadrillas': cuadrillas
    })


@role_required('admin', 'bodega')
def crear_cuadrilla(request):

    form = CrearCuadrillaConResponsableForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('/cuadrillas/')

    return render(request, 'cuadrillas/crear_cuadrilla.html', {
        'form': form
    })


@role_required('admin', 'bodega')
def importar_clientes(request):

    if request.method == 'POST':
        archivo = request.FILES.get('archivo_clientes')

        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo CSV o XLSX.')
            return redirect('importar_clientes')

        try:
            clientes_archivo = _leer_clientes_archivo(archivo)
        except Exception as exc:
            messages.error(request, f'No se pudo leer el archivo: {exc}')
            return redirect('importar_clientes')

        vistos = set()
        creados = 0
        duplicados = 0
        actualizados = 0

        for fila in clientes_archivo:
            nombre = fila.get('nombre', '')
            ct = fila.get('ct', '')
            cedula = fila.get('cedula', '')
            clave = nombre.lower()
            if clave in vistos:
                duplicados += 1
                continue
            vistos.add(clave)

            cliente = Cliente.objects.filter(nombre__iexact=nombre).first()
            if cliente:
                cambios = []
                if ct and cliente.ct != ct:
                    cliente.ct = ct
                    cambios.append('ct')
                if cedula and cliente.cedula != cedula:
                    cliente.cedula = cedula
                    cambios.append('cedula')

                if cambios:
                    cliente.save(update_fields=cambios)
                    actualizados += 1
                else:
                    duplicados += 1
            else:
                Cliente.objects.create(nombre=nombre, ct=ct, cedula=cedula)
                creados += 1

        messages.success(
            request,
            f'Importacion de clientes finalizada. Creados: {creados}, Actualizados: {actualizados}, Duplicados: {duplicados}.'
        )
        return redirect('importar_clientes')

    return render(request, 'cuadrillas/importar_clientes.html', {
        'total_clientes': Cliente.objects.count(),
    })


@role_required('admin', 'bodega')
def editar_cuadrilla(request, id):

    cuadrilla = get_object_or_404(Cuadrilla, id=id)

    form = CuadrillaForm(request.POST or None, instance=cuadrilla)

    if form.is_valid():
        form.save()
        return redirect('/cuadrillas/')

    return render(request, 'cuadrillas/editar_cuadrilla.html', {
        'form': form
    })


@role_required('admin', 'bodega')
def eliminar_cuadrilla(request, id):

    cuadrilla = get_object_or_404(Cuadrilla, id=id)

    cuadrilla.delete()

    return redirect('/cuadrillas/')


@role_required('admin')
def lista_soportes(request):

    soportes = SoporteTecnico.objects.all()

    return render(request, 'cuadrillas/soportes.html', {
        'soportes': soportes
    })


@role_required('admin')
def crear_soporte(request):

    form = CrearSoporteConResponsableForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('/soportes/')

    return render(request, 'cuadrillas/crear_soporte.html', {
        'form': form
    })


@role_required('admin')
def editar_soporte(request, id):

    soporte = get_object_or_404(SoporteTecnico, id=id)

    form = SoporteTecnicoForm(request.POST or None, instance=soporte)

    if form.is_valid():
        form.save()
        return redirect('/soportes/')

    return render(request, 'cuadrillas/editar_soporte.html', {
        'form': form
    })


@role_required('admin')
def eliminar_soporte(request, id):

    soporte = get_object_or_404(SoporteTecnico, id=id)

    soporte.delete()

    return redirect('/soportes/')

@role_required('admin', 'bodega', 'soporte')
def asignar_material(request):
    def _identificadores_disponibles(producto, es_activo=False, es_fibra=False, es_codigo_unico=False):
        tipos = []
        if es_activo:
            tipos = ['SERIAL', 'MAC']
        elif es_fibra:
            tipos = ['CABLE']
        elif es_codigo_unico:
            tipos = ['SERIAL']

        qs = DetalleIdentificador.objects.filter(
            producto=producto,
            tipo__in=tipos
        ).filter(
            asignacion_detalle__isnull=True
        )

        if es_activo:
            identificadores_en_baja = BajaIdentificador.objects.filter(
                identificador__producto=producto
            ).values_list('identificador_id', flat=True)
            qs = qs.exclude(id__in=identificadores_en_baja)

        if es_fibra:
            qs = qs.filter(metraje__gt=0)

        return qs.order_by('tipo', 'valor')

    cuadrillas = Cuadrilla.objects.select_related('responsable').order_by('nombre')
    productos = Producto.objects.select_related('categoria', 'marca', 'modelo', 'tipo_fibra').filter(
        Q(stock__gt=0) | Q(metraje__gt=0)
    ).order_by('nombre')

    productos_config = {}
    for producto in productos:
        categoria_nombre = producto.categoria.nombre if producto.categoria else ''
        marca_nombre = producto.marca.nombre if producto.marca else ''
        modelo_nombre = producto.modelo.nombre if producto.modelo else ''
        activo = es_categoria_activa(categoria_nombre)
        fibra = _producto_es_fibra(producto)
        codigo_unico = es_categoria_codigo_unico(categoria_nombre)
        identificadores = []

        if activo or fibra or codigo_unico:
            for item in _identificadores_disponibles(
                producto,
                es_activo=activo,
                es_fibra=fibra,
                es_codigo_unico=codigo_unico,
            ):
                codigo_item = (item.codigo_individual or '').strip()
                etiqueta = f'{item.tipo}: {item.valor}'
                if codigo_item:
                    etiqueta += f'  Codigo: {codigo_item}'
                identificadores.append({
                    'id': item.id,
                    'etiqueta': etiqueta,
                    'metraje': float(item.metraje or 0),
                })

        productos_config[str(producto.id)] = {
            'nombre': producto.nombre,
            'codigo': producto.codigo,
            'categoria': categoria_nombre,
            'marca': marca_nombre,
            'modelo': modelo_nombre,
            'stock': int(producto.stock or 0),
            'metraje': float(producto.metraje or 0),
            'es_activo': activo,
            'es_fibra': fibra,
            'es_codigo_unico': codigo_unico,
            'identificadores': identificadores,
        }

    producto_preseleccionado = (request.GET.get('producto') or '').strip()

    if request.method == 'POST':
        cuadrilla_id = (request.POST.get('cuadrilla') or '').strip()
        producto_ids = request.POST.getlist('producto[]')
        cantidades = request.POST.getlist('cantidad[]')
        metrajes = request.POST.getlist('metraje[]')
        identificadores_por_linea = request.POST.getlist('identificadores[]')

        if not cuadrilla_id:
            messages.error(request, 'Debe seleccionar una cuadrilla.')
        else:
            cuadrilla = Cuadrilla.objects.filter(id=cuadrilla_id).first()
            if not cuadrilla:
                messages.error(request, 'La cuadrilla seleccionada no existe.')
            else:
                lineas = []
                errores = []
                acumulado_stock = defaultdict(int)
                acumulado_metraje = defaultdict(Decimal)
                identificadores_tomados = set()

                for idx, producto_id in enumerate(producto_ids):
                    producto_id = (producto_id or '').strip()
                    if not producto_id:
                        continue

                    producto = Producto.objects.select_related('categoria', 'tipo_fibra').filter(id=producto_id).first()
                    if not producto:
                        errores.append(f'Linea {idx + 1}: el producto seleccionado no existe.')
                        continue

                    try:
                        cantidad = int((cantidades[idx] if idx < len(cantidades) else '0') or '0')
                    except (TypeError, ValueError):
                        cantidad = 0

                    if cantidad <= 0:
                        errores.append(f'Linea {idx + 1} ({producto.nombre}): la cantidad debe ser mayor a cero.')
                        continue

                    categoria_nombre = producto.categoria.nombre if producto.categoria else ''
                    es_activo = es_categoria_activa(categoria_nombre)
                    es_fibra = _producto_es_fibra(producto)
                    es_codigo_unico = es_categoria_codigo_unico(categoria_nombre)
                    metraje = _parse_decimal_input(metrajes[idx] if idx < len(metrajes) else '0')

                    if es_fibra and metraje <= 0:
                        errores.append(f'Linea {idx + 1} ({producto.nombre}): debe ingresar metraje mayor a cero.')
                        continue

                    if not es_fibra:
                        metraje = Decimal('0')

                    ids_crudos = (identificadores_por_linea[idx] if idx < len(identificadores_por_linea) else '').strip()
                    ids_seleccionados = [x.strip() for x in ids_crudos.split(',') if x.strip()]

                    if es_activo or es_fibra or es_codigo_unico:
                        if len(ids_seleccionados) != cantidad:
                            etiqueta = 'seriales/MAC' if es_activo else ('codigos de bobina/cable' if es_fibra else 'codigos unicos')
                            errores.append(f'Linea {idx + 1} ({producto.nombre}): la cantidad debe coincidir con {etiqueta} seleccionados.')
                            continue

                        disponibles = set(
                            _identificadores_disponibles(
                                producto,
                                es_activo=es_activo,
                                es_fibra=es_fibra,
                                es_codigo_unico=es_codigo_unico,
                            ).values_list('id', flat=True)
                        )
                        ids_enteros = []
                        ids_invalidos = False
                        for id_item in ids_seleccionados:
                            try:
                                id_int = int(id_item)
                            except ValueError:
                                ids_invalidos = True
                                break
                            ids_enteros.append(id_int)

                        if ids_invalidos:
                            errores.append(f'Linea {idx + 1} ({producto.nombre}): identificadores invalidos.')
                            continue

                        repetidos_en_linea = len(set(ids_enteros)) != len(ids_enteros)
                        if repetidos_en_linea:
                            errores.append(f'Linea {idx + 1} ({producto.nombre}): hay identificadores repetidos en la misma linea.')
                            continue

                        if not set(ids_enteros).issubset(disponibles):
                            errores.append(f'Linea {idx + 1} ({producto.nombre}): hay identificadores no disponibles.')
                            continue

                        if identificadores_tomados.intersection(set(ids_enteros)):
                            errores.append(f'Linea {idx + 1} ({producto.nombre}): hay identificadores repetidos entre lineas.')
                            continue

                        identificadores_qs = DetalleIdentificador.objects.filter(id__in=ids_enteros).only('id', 'metraje')
                        metraje_calculado = sum((item.metraje or Decimal('0')) for item in identificadores_qs)

                        if es_fibra:
                            if metraje_calculado <= 0:
                                errores.append(f'Linea {idx + 1} ({producto.nombre}): la bobina seleccionada no tiene metraje disponible.')
                                continue
                            metraje = metraje_calculado

                        identificadores_tomados.update(set(ids_enteros))
                        ids_seleccionados = ids_enteros
                    else:
                        ids_seleccionados = []

                    acumulado_stock[producto.id] += cantidad
                    if es_fibra:
                        acumulado_metraje[producto.id] += metraje

                    lineas.append({
                        'producto': producto,
                        'cantidad': cantidad,
                        'metraje': metraje,
                        'es_fibra': es_fibra,
                        'es_activo': es_activo,
                        'es_codigo_unico': es_codigo_unico,
                        'identificadores': ids_seleccionados,
                    })

                if not lineas and not errores:
                    errores.append('Debe agregar al menos una linea con producto y cantidad.')

                for linea in lineas:
                    producto = linea['producto']
                    requerido_stock = acumulado_stock[producto.id]
                    requerido_metraje = acumulado_metraje.get(producto.id, Decimal('0'))

                    if (not linea['es_fibra']) and producto.stock < requerido_stock:
                        errores.append(
                            f'No hay suficiente stock para {producto.nombre}. Disponible: {producto.stock}, solicitado: {requerido_stock}.'
                        )

                    if linea['es_fibra']:
                        metraje_real_disponible = DetalleIdentificador.objects.filter(
                            producto=producto,
                            tipo='CABLE',
                            metraje__gt=0,
                            asignacion_detalle__isnull=True,
                        ).aggregate(total=Sum('metraje'))['total'] or Decimal('0')
                        if metraje_real_disponible < requerido_metraje:
                            errores.append(
                                f'No hay suficiente metraje para {producto.nombre}. Disponible: {metraje_real_disponible}, solicitado: {requerido_metraje}.'
                            )

                if errores:
                    for err in errores:
                        messages.error(request, err)
                else:
                    with transaction.atomic():
                        usuario_registro = request.user.get_username()
                        for linea in lineas:
                            producto = linea['producto']
                            cantidad = linea['cantidad']
                            metraje = linea['metraje']

                            existente = AsignacionMaterial.objects.filter(
                                producto=producto,
                                cuadrilla=cuadrilla
                            ).first()

                            if existente:
                                existente.cantidad += cantidad
                                if linea['es_fibra']:
                                    existente.metraje += metraje
                                existente.usuario_registro = usuario_registro
                                existente.fecha = timezone.localdate()
                                existente.save(update_fields=['cantidad', 'metraje', 'usuario_registro', 'fecha'])
                                destino = existente
                            else:
                                destino = AsignacionMaterial.objects.create(
                                    producto=producto,
                                    cuadrilla=cuadrilla,
                                    cantidad=cantidad,
                                    metraje=metraje if linea['es_fibra'] else Decimal('0'),
                                    usuario_registro=usuario_registro,
                                )

                            if not linea['es_fibra']:
                                producto.stock -= cantidad
                            if linea['es_fibra']:
                                producto.metraje -= metraje
                            producto.save()

                            if linea['es_activo'] or linea['es_fibra'] or linea['es_codigo_unico']:
                                identificadores = DetalleIdentificador.objects.filter(
                                    id__in=linea['identificadores']
                                )
                                for identificador in identificadores:
                                    AsignacionIdentificador.objects.create(
                                        asignacion=destino,
                                        identificador=identificador
                                    )

                    messages.success(request, 'Materiales asignados correctamente en un solo envio.')
                    return redirect('lista_asignaciones')

    return render(
        request,
        'cuadrillas/asignar_material.html',
        {
            'cuadrillas': cuadrillas,
            'productos': productos,
            'productos_config_json': json.dumps(productos_config),
            'producto_preseleccionado': producto_preseleccionado,
        }
    )


@role_required('admin', 'bodega', 'cuadrilla', 'soporte')
def lista_asignaciones(request):

    usuario = request.user

    #FILTRO SEGÚN ROL
    if _usuario_es_cuadrilla(usuario):
        filtro = _filtro_cuadrilla_por_usuario(usuario)
        asignaciones = AsignacionMaterial.objects.select_related(
            'producto', 'cuadrilla'
        ).prefetch_related(
            'identificadores_asignados__identificador'
        ).filter(filtro, cantidad__gt=0)

    else:
        asignaciones = AsignacionMaterial.objects.select_related(
            'producto', 'cuadrilla'
        ).prefetch_related(
            'identificadores_asignados__identificador'
        ).filter(cantidad__gt=0)

    # AGRUPACIÓN POR CUADRILLA (sin consolidar filas, para poder dar de baja por asignación)
    cuadrillas = defaultdict(list)
    for a in asignaciones:
        detalles = list(a.identificadores_asignados.select_related('identificador').all())
        es_activo = es_categoria_activa(a.producto.categoria.nombre if a.producto.categoria else '')
        es_fibra = _producto_es_fibra(a.producto)
        es_adss = _producto_es_adss(a.producto)
        es_codigo_unico = es_categoria_codigo_unico(a.producto.categoria.nombre if a.producto.categoria else '')
        es_construccion = a.producto.bodega == Producto.BODEGA_CONSTRUCCION

        if es_fibra:
            detalles = [
                d for d in detalles
                if d.identificador.tipo == 'CABLE' and (d.identificador.metraje or Decimal('0')) > 0
            ]
        elif es_activo:
            detalles = [d for d in detalles if d.identificador.tipo in ['SERIAL', 'MAC']]
        elif es_codigo_unico:
            detalles = [d for d in detalles if d.identificador.tipo == 'SERIAL']

        cuadrillas[a.cuadrilla.nombre].append({
            'id': a.id,
            'cuadrilla': a.cuadrilla.nombre,
            'producto': a.producto.nombre,
            'codigo': a.producto.codigo,
            'marca': a.producto.marca,
            'cantidad': a.cantidad,
            'metraje': a.metraje,
            'usuario_registro': a.usuario_registro,
            'fecha': a.fecha,
            'es_activo': es_activo,
            'es_fibra': es_fibra,
            'es_adss': es_adss,
            'es_codigo_unico': es_codigo_unico,
            'es_construccion': es_construccion,
            'identificadores': [
                {
                    'id': d.id,
                    'valor': d.identificador.valor,
                    'tipo': d.identificador.tipo,
                    'codigo_individual': d.identificador.codigo_individual,
                    'metraje': d.identificador.metraje,
                }
                for d in detalles
            ],
        })

    es_admin = usuario.is_superuser or usuario.is_staff

    categorias_activas = [
        categoria
        for categoria in Categoria.objects.order_by('nombre')
        if es_categoria_activa(categoria.nombre)
    ]

    modelos = list(
        ModeloEquipo.objects.select_related('marca')
        .order_by('nombre')
        .values('id', 'nombre', 'marca_id')
    )

    seriales_actuales = list(
        DetalleIdentificador.objects.filter(tipo__in=['SERIAL', 'MAC'])
        .exclude(valor='')
        .values_list('valor', flat=True)[:1500]
    )
    seriales_retirados = list(
        BajaMaterial.objects.exclude(serial_mac_retirado='')
        .values_list('serial_mac_retirado', flat=True)[:1500]
    )
    serial_mac_sugeridos = sorted({
        str(valor).strip().upper()
        for valor in [*seriales_actuales, *seriales_retirados]
        if str(valor or '').strip()
    })

    return render(request, 'cuadrillas/lista_asignaciones.html', {
        'cuadrillas': dict(cuadrillas),
        'clientes': Cliente.objects.order_by('nombre'),
        'es_admin': es_admin,
        'categorias': categorias_activas,
        'marcas': Marca.objects.order_by('nombre'),
        'modelos_json': json.dumps(modelos),
        'serial_mac_sugeridos': serial_mac_sugeridos,
    })


@role_required('admin', 'bodega', 'cuadrilla', 'soporte')
def dar_baja_material(request, asignacion_id):
    if request.method != 'POST':
        return redirect('lista_asignaciones')

    asignacion = get_object_or_404(AsignacionMaterial, id=asignacion_id)

    if _usuario_es_cuadrilla(request.user):
        filtro = _filtro_cuadrilla_por_usuario(request.user)
        permitido = AsignacionMaterial.objects.filter(id=asignacion.id).filter(filtro).exists()
        if not permitido:
            messages.error(request, 'No tiene permisos para dar de baja esta asignación.')
            return redirect('lista_asignaciones')

    utilizado_en_opcion = (request.POST.get('utilizado_en_opcion') or '').strip()
    cliente_nombre = _normalizar_cliente_nombre(request.POST.get('cliente_nombre'))
    observacion = (request.POST.get('observacion_cambio') or '').strip()
    serial_mac_retirado = (request.POST.get('serial_mac_retirado') or '').strip().upper()
    codigo_retirado = (request.POST.get('codigo_retirado') or '').strip()
    categoria_retirado = (request.POST.get('categoria_retirado') or '').strip()
    marca_retirado = (request.POST.get('marca_retirado') or '').strip()
    modelo_retirado = (request.POST.get('modelo_retirado') or '').strip()
    coordenadas_baja = (request.POST.get('coordenadas_baja') or '').strip()
    sector_baja = (request.POST.get('sector_baja') or '').strip()
    coord_cliente = (request.POST.get('coord_cliente') or '').strip()
    coord_caja = (request.POST.get('coord_caja') or '').strip()
    coord_punta_inicial = (request.POST.get('coord_punta_inicial') or '').strip()
    coord_punta_final = (request.POST.get('coord_punta_final') or '').strip()
    activo = es_categoria_activa(asignacion.producto.categoria.nombre if asignacion.producto.categoria else '')
    es_fibra = _producto_es_fibra(asignacion.producto)
    es_adss = _producto_es_adss(asignacion.producto)
    es_codigo_unico = es_categoria_codigo_unico(asignacion.producto.categoria.nombre if asignacion.producto.categoria else '')
    es_construccion = asignacion.producto.bodega == Producto.BODEGA_CONSTRUCCION

    if not utilizado_en_opcion:
        messages.error(request, 'Debe seleccionar una opcion en "Utilizado en".')
        return redirect('lista_asignaciones')

    if (not es_construccion) and (not cliente_nombre):
        messages.error(request, 'Debe seleccionar o escribir un cliente.')
        return redirect('lista_asignaciones')

    if activo and utilizado_en_opcion == 'Cambio' and not observacion:
        messages.error(request, 'Debe ingresar el detalle del cambio.')
        return redirect('lista_asignaciones')

    if activo and utilizado_en_opcion == 'Cambio' and not serial_mac_retirado:
        messages.error(request, 'Debe ingresar el Serial/MAC del equipo retirado para cambios.')
        return redirect('lista_asignaciones')

    if es_construccion:
        if utilizado_en_opcion not in ['DANO', 'CONSTRUCCION']:
            messages.error(request, 'Para bodega de construccion solo se permite "DAÑO" o "CONSTRUCCION" en Utilizado en.')
            return redirect('lista_asignaciones')
        if not coordenadas_baja or not sector_baja:
            messages.error(request, 'Debe ingresar Coordenadas y Sector para la baja de construccion.')
            return redirect('lista_asignaciones')
        partes_observacion = [
            f'Coordenadas: {coordenadas_baja}',
            f'Sector: {sector_baja}',
        ]
        if es_adss and es_fibra:
            if not coord_punta_inicial or not coord_punta_final:
                messages.error(request, 'Para ADSS debe ingresar coordenada de punta inicial y coordenada de punta final.')
                return redirect('lista_asignaciones')
            partes_observacion.append(f'Coord punta inicial: {coord_punta_inicial}')
            partes_observacion.append(f'Coord punta final: {coord_punta_final}')
        observacion = ' | '.join(partes_observacion)
    elif es_adss and es_fibra:
        if not coord_punta_inicial or not coord_punta_final:
            messages.error(request, 'Para ADSS debe ingresar coordenada de punta inicial y coordenada de punta final.')
            return redirect('lista_asignaciones')
        coordenadas_punta = [
            f'Coord punta inicial: {coord_punta_inicial}',
            f'Coord punta final: {coord_punta_final}',
        ]
        observacion = f"{observacion} | {' | '.join(coordenadas_punta)}" if observacion else ' | '.join(coordenadas_punta)

    if es_fibra or activo:
        coordenadas_extra = []
        if coord_cliente:
            coordenadas_extra.append(f'Coord cliente: {coord_cliente}')
        if coord_caja:
            coordenadas_extra.append(f'Coord caja: {coord_caja}')
        if coordenadas_extra:
            bloque = ' | '.join(coordenadas_extra)
            observacion = f"{observacion} | {bloque}" if observacion else bloque

    cliente = None
    if cliente_nombre:
        cliente = Cliente.objects.filter(nombre__iexact=cliente_nombre).first()
        if not cliente:
            cliente = Cliente.objects.create(nombre=cliente_nombre)

    punta_inicial = Decimal('0')
    punta_final = Decimal('0')
    codigo_bobina = (request.POST.get('codigo_bobina') or '').strip()

    if es_fibra:
        punta_inicial = _parse_decimal_input(request.POST.get('punta_inicial'))
        punta_final = _parse_decimal_input(request.POST.get('punta_final'))
        metraje_baja = _parse_decimal_input(request.POST.get('metraje_baja'))
        cantidad_baja = 1

        if punta_inicial <= 0 and punta_final <= 0:
            messages.error(request, 'Debe ingresar punta inicial y punta final para bajas de fibra.')
            return redirect('lista_asignaciones')

        if punta_final == punta_inicial:
            messages.error(request, 'La punta inicial y la punta final no pueden ser iguales.')
            return redirect('lista_asignaciones')

        metraje_calculado = abs(punta_final - punta_inicial)
        if metraje_baja <= 0:
            metraje_baja = metraje_calculado

        diferencia = abs(metraje_baja - metraje_calculado)
        if diferencia > Decimal('0.01'):
            messages.error(request, 'El total debe coincidir con la diferencia entre punta final y punta inicial.')
            return redirect('lista_asignaciones')

        if metraje_baja <= 0:
            messages.error(request, 'El metraje a dar de baja debe ser mayor a cero.')
            return redirect('lista_asignaciones')

        if metraje_baja > (asignacion.metraje or Decimal('0')):
            messages.error(request, 'El metraje de baja no puede ser mayor al metraje asignado.')
            return redirect('lista_asignaciones')
    else:
        try:
            cantidad_baja = int(request.POST.get('cantidad_baja', '0'))
        except ValueError:
            cantidad_baja = 0

        metraje_baja = Decimal('0')

        if cantidad_baja <= 0:
            messages.error(request, 'La cantidad a dar de baja debe ser mayor a cero.')
            return redirect('lista_asignaciones')

        if cantidad_baja > asignacion.cantidad:
            messages.error(request, 'La cantidad de baja no puede ser mayor a la cantidad asignada.')
            return redirect('lista_asignaciones')

    if activo:
        ids_detalle = request.POST.getlist('identificadores_baja')
        if len(ids_detalle) != cantidad_baja:
            messages.error(request, 'Para equipos activos, seleccione un serial por cada unidad dada de baja.')
            return redirect('lista_asignaciones')

        detalles_qs = AsignacionIdentificador.objects.filter(
            id__in=ids_detalle,
            asignacion=asignacion
        )
        if detalles_qs.count() != cantidad_baja:
            messages.error(request, 'Algunos seriales seleccionados no corresponden a esta asignación.')
            return redirect('lista_asignaciones')

        detalles_para_baja = list(detalles_qs.select_related('identificador'))

        for detalle in detalles_para_baja:
            if detalle.identificador.producto_id != asignacion.producto_id:
                messages.error(request, 'El identificador seleccionado no corresponde al producto de la asignacion.')
                return redirect('lista_asignaciones')
            if detalle.identificador.tipo not in ['SERIAL', 'MAC']:
                messages.error(request, 'Para equipos activos solo se permite registrar Serial o MAC en la baja.')
                return redirect('lista_asignaciones')
    elif es_codigo_unico:
        ids_detalle = request.POST.getlist('identificadores_baja')
        if len(ids_detalle) != cantidad_baja:
            messages.error(request, 'Para MANGAS/CAJAS, seleccione un codigo unico por cada unidad dada de baja.')
            return redirect('lista_asignaciones')

        detalles_qs = AsignacionIdentificador.objects.filter(
            id__in=ids_detalle,
            asignacion=asignacion
        )
        if detalles_qs.count() != cantidad_baja:
            messages.error(request, 'Algunos codigos seleccionados no corresponden a esta asignacion.')
            return redirect('lista_asignaciones')

        detalles_para_baja = list(detalles_qs.select_related('identificador'))
        for detalle in detalles_para_baja:
            if detalle.identificador.producto_id != asignacion.producto_id:
                messages.error(request, 'El identificador seleccionado no corresponde al producto de la asignacion.')
                return redirect('lista_asignaciones')
            if detalle.identificador.tipo != 'SERIAL':
                messages.error(request, 'Para MANGAS/CAJAS solo se permite codigo unico (SERIAL).')
                return redirect('lista_asignaciones')
    elif es_fibra:
        detalles_fibra = list(
            AsignacionIdentificador.objects.filter(asignacion=asignacion).select_related('identificador')
        )
        if detalles_fibra:
            ids_detalle = request.POST.getlist('identificadores_baja')
            detalle_encontrado = None

            if ids_detalle:
                if len(ids_detalle) != 1:
                    messages.error(request, 'Seleccione un solo identificador de fibra/utp/coaxial para la baja.')
                    return redirect('lista_asignaciones')
                detalle_encontrado = AsignacionIdentificador.objects.filter(
                    id=ids_detalle[0],
                    asignacion=asignacion
                ).select_related('identificador').first()
                if not detalle_encontrado:
                    messages.error(request, 'El identificador seleccionado no corresponde a esta asignacion.')
                    return redirect('lista_asignaciones')
                codigo_bobina = (
                    (detalle_encontrado.identificador.codigo_individual or '').strip()
                    or (detalle_encontrado.identificador.valor or '').strip()
                )
            else:
                if not codigo_bobina:
                    messages.error(request, 'Debe seleccionar un identificador o ingresar codigo de bobina para la baja de fibra/cable.')
                    return redirect('lista_asignaciones')

                codigo_busqueda = codigo_bobina.lower()
                for detalle in detalles_fibra:
                    codigo_item = (detalle.identificador.codigo_individual or '').strip().lower()
                    valor_item = (detalle.identificador.valor or '').strip().lower()
                    if codigo_busqueda in [codigo_item, valor_item]:
                        detalle_encontrado = detalle
                        break

                if not detalle_encontrado:
                    messages.error(request, 'El codigo de bobina no corresponde a la asignacion seleccionada.')
                    return redirect('lista_asignaciones')

            metraje_identificador = detalle_encontrado.identificador.metraje or Decimal('0')
            if metraje_identificador > 0 and metraje_baja > metraje_identificador:
                messages.error(
                    request,
                    f'La baja ({metraje_baja} m) no puede ser mayor al metraje disponible de la bobina seleccionada ({metraje_identificador} m).'
                )
                return redirect('lista_asignaciones')

            detalles_para_baja = [detalle_encontrado]
        else:
            if not codigo_bobina:
                messages.error(request, 'Debe ingresar codigo de bobina para la baja de fibra/cable.')
                return redirect('lista_asignaciones')
            detalles_para_baja = []
    else:
        detalles_para_baja = []

    lista_mac = [
        d.identificador.valor
        for d in detalles_para_baja
        if d.identificador.tipo == 'MAC'
    ]
    lista_serial = [
        d.identificador.valor
        for d in detalles_para_baja
        if d.identificador.tipo == 'SERIAL'
    ]

    codigos_equipo = [
        (d.identificador.codigo_individual or '').strip()
        for d in detalles_para_baja
        if (d.identificador.codigo_individual or '').strip()
    ]
    codigo_baja = codigo_bobina if es_fibra else (', '.join(codigos_equipo) if codigos_equipo else asignacion.producto.codigo)

    baja = BajaMaterial.objects.create(
        asignacion=asignacion,
        cantidad=cantidad_baja,
        metraje=metraje_baja,
        punta_inicial=punta_inicial,
        punta_final=punta_final,
        usuario_registro=request.user.get_username(),
        tipo_uso=utilizado_en_opcion,
        cliente=cliente,
        cliente_nombre=cliente.nombre if cliente else '',
        producto_nombre=asignacion.producto.nombre,
        producto_codigo=codigo_baja,
        codigo_bobina=codigo_bobina if es_fibra else '',
        es_equipo_activo=activo,
        observacion=observacion,
        serial_mac_retirado=serial_mac_retirado if utilizado_en_opcion == 'Cambio' else '',
        codigo_retirado=codigo_retirado if utilizado_en_opcion == 'Cambio' else '',
        categoria_retirado=categoria_retirado if utilizado_en_opcion == 'Cambio' else '',
        marca_retirado=marca_retirado if utilizado_en_opcion == 'Cambio' else '',
        modelo_retirado=modelo_retirado if utilizado_en_opcion == 'Cambio' else '',
        estado_equipo=(request.POST.get('estado_equipo') or '') if (activo or es_codigo_unico) else '',
        detalle_mac=', '.join(lista_mac),
        detalle_serial=', '.join(lista_serial),
    )

    if activo and utilizado_en_opcion == 'Cambio' and serial_mac_retirado:
        messages.info(request, 'Equipo retirado guardado en bodega de retirados. Puede reutilizarlo desde la seccion Equipos Retirados.')

    for detalle in detalles_para_baja:
        BajaIdentificador.objects.create(
            baja=baja,
            identificador=detalle.identificador,
        )

    if detalles_para_baja and not es_fibra:
        AsignacionIdentificador.objects.filter(id__in=[d.id for d in detalles_para_baja]).delete()

    if es_fibra and detalles_para_baja:
        detalle_fibra = detalles_para_baja[0]
        identificador_fibra = detalle_fibra.identificador
        metraje_actual = identificador_fibra.metraje or Decimal('0')
        nuevo_metraje = metraje_actual - metraje_baja
        if nuevo_metraje <= Decimal('0.01'):
            identificador_fibra.metraje = Decimal('0')
            identificador_fibra.save(update_fields=['metraje'])
            AsignacionIdentificador.objects.filter(id=detalle_fibra.id).delete()
        else:
            identificador_fibra.metraje = nuevo_metraje
            identificador_fibra.save(update_fields=['metraje'])

    if es_fibra:
        asignacion.metraje -= metraje_baja
        if asignacion.metraje < 0:
            asignacion.metraje = Decimal('0')
        if asignacion.metraje == 0:
            asignacion.cantidad = 0
    else:
        asignacion.cantidad -= cantidad_baja
        if asignacion.cantidad < 0:
            asignacion.cantidad = 0
    asignacion.save()

    messages.success(request, 'Material dado de baja correctamente.')
    return redirect('lista_asignaciones')


@role_required('admin')
def devolver_a_stock(request, asignacion_id):
    if request.method != 'POST':
        return redirect('lista_asignaciones')

    asignacion = get_object_or_404(AsignacionMaterial, id=asignacion_id)

    producto = asignacion.producto
    activo = es_categoria_activa(producto.categoria.nombre if producto.categoria else '')
    es_fibra = _producto_es_fibra(producto)

    if activo:
        ids_detalle = request.POST.getlist('identificadores_devolver')
        if not ids_detalle:
            messages.error(request, 'Debe seleccionar al menos un serial/MAC para devolver a stock.')
            return redirect('lista_asignaciones')

        detalles_qs = AsignacionIdentificador.objects.filter(
            id__in=ids_detalle,
            asignacion=asignacion,
        ).select_related('identificador')

        if detalles_qs.count() != len(ids_detalle):
            messages.error(request, 'Hay seriales/MAC que no pertenecen a esta asignación.')
            return redirect('lista_asignaciones')

        cantidad_devolver = detalles_qs.count()
        if cantidad_devolver > asignacion.cantidad:
            messages.error(request, 'La devolución supera la cantidad asignada.')
            return redirect('lista_asignaciones')

        with transaction.atomic():
            AsignacionIdentificador.objects.filter(id__in=[d.id for d in detalles_qs]).delete()
            asignacion.cantidad -= cantidad_devolver
            if asignacion.cantidad < 0:
                asignacion.cantidad = 0
            asignacion.save(update_fields=['cantidad'])

            producto.stock = (producto.stock or 0) + cantidad_devolver
            producto.save(update_fields=['stock'])

        messages.success(request, f'Se devolvieron {cantidad_devolver} serial(es)/MAC a stock correctamente.')
        return redirect('lista_asignaciones')

    if es_fibra:
        detalle_fibra = None
        detalles_fibra_asignados = list(
            AsignacionIdentificador.objects.filter(asignacion=asignacion)
            .select_related('identificador')
            .filter(identificador__tipo='CABLE')
        )

        id_detalle_fibra = (request.POST.get('identificador_fibra_devolver') or '').strip()
        if detalles_fibra_asignados and not id_detalle_fibra:
            messages.error(request, 'Debe seleccionar una bobina/cable para devolver a stock.')
            return redirect('lista_asignaciones')

        if id_detalle_fibra:
            detalle_fibra = AsignacionIdentificador.objects.filter(
                id=id_detalle_fibra,
                asignacion=asignacion,
                identificador__tipo='CABLE'
            ).select_related('identificador').first()
            if not detalle_fibra:
                messages.error(request, 'La bobina seleccionada no corresponde a esta asignación.')
                return redirect('lista_asignaciones')

        metraje_devolver = _parse_decimal_input(request.POST.get('metraje_devolver'))
        if metraje_devolver <= 0:
            messages.error(request, 'El metraje a devolver debe ser mayor a cero.')
            return redirect('lista_asignaciones')
        if metraje_devolver > (asignacion.metraje or Decimal('0')):
            messages.error(request, 'El metraje a devolver no puede ser mayor al metraje asignado.')
            return redirect('lista_asignaciones')

        if detalle_fibra:
            metraje_disponible_bobina = detalle_fibra.identificador.metraje or Decimal('0')
            if metraje_disponible_bobina <= 0:
                messages.error(request, 'La bobina seleccionada no tiene metraje disponible para devolver.')
                return redirect('lista_asignaciones')
            if metraje_devolver > metraje_disponible_bobina:
                messages.error(
                    request,
                    f'El metraje a devolver ({metraje_devolver} m) no puede ser mayor al metraje disponible de la bobina ({metraje_disponible_bobina} m).'
                )
                return redirect('lista_asignaciones')

            # En devolucion a stock se retorna la bobina completa disponible.
            if abs(metraje_devolver - metraje_disponible_bobina) > Decimal('0.01'):
                messages.error(
                    request,
                    f'Debe devolver el metraje completo de la bobina seleccionada ({metraje_disponible_bobina} m).'
                )
                return redirect('lista_asignaciones')

        with transaction.atomic():
            asignacion.metraje = (asignacion.metraje or Decimal('0')) - metraje_devolver
            if asignacion.metraje < 0:
                asignacion.metraje = Decimal('0')

            if detalle_fibra:
                AsignacionIdentificador.objects.filter(id=detalle_fibra.id).delete()
                asignacion.cantidad = max((asignacion.cantidad or 0) - 1, 0)

            if asignacion.metraje == 0:
                asignacion.cantidad = 0
            asignacion.save(update_fields=['metraje', 'cantidad'])

            producto.metraje = (producto.metraje or Decimal('0')) + metraje_devolver
            producto.save(update_fields=['metraje'])

        messages.success(request, f'Se devolvieron {metraje_devolver} m a stock correctamente.')
        return redirect('lista_asignaciones')

    try:
        cantidad_devolver = int(request.POST.get('cantidad_devolver', '0'))
    except ValueError:
        cantidad_devolver = 0

    if cantidad_devolver <= 0:
        messages.error(request, 'La cantidad a devolver debe ser mayor a cero.')
        return redirect('lista_asignaciones')
    if cantidad_devolver > asignacion.cantidad:
        messages.error(request, 'La cantidad a devolver no puede ser mayor a la cantidad asignada.')
        return redirect('lista_asignaciones')

    with transaction.atomic():
        asignacion.cantidad -= cantidad_devolver
        if asignacion.cantidad < 0:
            asignacion.cantidad = 0
        asignacion.save(update_fields=['cantidad'])

        producto.stock = (producto.stock or 0) + cantidad_devolver
        producto.save(update_fields=['stock'])

    messages.success(request, f'Se devolvieron {cantidad_devolver} unidad(es) a stock correctamente.')
    return redirect('lista_asignaciones')


@role_required('admin', 'bodega')
def historial_bajas(request):
    usuario = request.user
    cliente_nombre = _normalizar_cliente_nombre(request.GET.get('cliente'))
    proveedor_id = request.GET.get('proveedor', '').strip()

    bajas_qs = BajaMaterial.objects.select_related(
        'cliente',
        'asignacion__cuadrilla',
        'asignacion__producto',
        'asignacion__producto__proveedor',
    ).prefetch_related(
        'identificadores_baja__identificador'
    ).exclude(
        asignacion__producto__bodega=Producto.BODEGA_CONSTRUCCION
    )

    if _usuario_es_cuadrilla(usuario):
        filtro = _filtro_cuadrilla_por_usuario(usuario)
        bajas_qs = bajas_qs.filter(filtro)

    if cliente_nombre:
        bajas_qs = bajas_qs.filter(
            Q(cliente__nombre__iexact=cliente_nombre) |
            Q(cliente_nombre__iexact=cliente_nombre)
        )

    if proveedor_id:
        bajas_qs = bajas_qs.filter(asignacion__producto__proveedor_id=proveedor_id)

    bajas = []
    for baja in bajas_qs.order_by('-fecha'):
        observacion_campos = _extraer_campos_observacion(baja.observacion)
        coordenadas = observacion_campos.get('coordenadas', '')
        sector = observacion_campos.get('sector', '')
        identificadores_baja = [
            f"{item.identificador.tipo}: {item.identificador.valor}"
            for item in baja.identificadores_baja.select_related('identificador').all()
        ]
        if baja.serial_mac_retirado:
            identificadores_baja.append(f"RETIRADO: {baja.serial_mac_retirado}")

        bajas.append({
            'id': baja.id,
            'fecha': baja.fecha,
            'cuadrilla': baja.asignacion.cuadrilla.nombre,
            'usuario_registro': baja.usuario_registro or '',
            'producto': baja.producto_nombre or baja.asignacion.producto.nombre,
            'codigo': baja.producto_codigo or baja.asignacion.producto.codigo,
            'cliente': baja.cliente_nombre or (baja.cliente.nombre if baja.cliente else ''),
            'tipo_uso': baja.tipo_uso,
            'observacion': _detalle_observacion_sin_coordenadas(baja.observacion),
            'coordenadas': coordenadas,
            'sector': sector,
            'coord_punta_inicial': observacion_campos.get('coord_punta_inicial', ''),
            'coord_punta_final': observacion_campos.get('coord_punta_final', ''),
            'coord_cliente': observacion_campos.get('coord_cliente', ''),
            'coord_caja': observacion_campos.get('coord_caja', ''),
            'cantidad': baja.cantidad,
            'punta_inicial': baja.punta_inicial,
            'punta_final': baja.punta_final,
            'codigo_bobina': baja.codigo_bobina,
            'metraje': baja.metraje,
            'proveedor': baja.asignacion.producto.proveedor.nombre if baja.asignacion.producto.proveedor else '-',
            'detalle_mac': baja.detalle_mac,
            'detalle_serial': baja.detalle_serial,
            'serial_mac_retirado': baja.serial_mac_retirado,
            'estado_equipo': baja.estado_equipo,
            'es_equipo_activo': baja.es_equipo_activo,
            'retirado': baja.retirado,
            'fecha_retirado': baja.fecha_retirado,
            'identificadores': identificadores_baja,
        })

    return render(request, 'cuadrillas/historial_bajas.html', {
        'bajas': bajas,
        'clientes': Cliente.objects.order_by('nombre'),
        'cliente_seleccionado': cliente_nombre,
        'proveedores': Proveedor.objects.order_by('nombre'),
        'proveedor_seleccionado': proveedor_id,
        'titulo_historial': 'Historial por Cliente',
        'limpiar_url_name': 'historial_bajas',
        'excel_url_name': 'reporte_historial_bajas_excel',
        'mostrar_excel': True,
        'mostrar_purgar': True,
        'purgar_url_name': 'purgar_historial_bajas',
        'mostrar_campos_construccion': False,
    })


@role_required('admin', 'bodega')
def historial_bajas_construccion(request):
    usuario = request.user
    bajas_qs = BajaMaterial.objects.select_related(
        'cliente',
        'asignacion__cuadrilla',
        'asignacion__producto',
        'asignacion__producto__proveedor',
    ).prefetch_related(
        'identificadores_baja__identificador'
    ).filter(
        asignacion__producto__bodega=Producto.BODEGA_CONSTRUCCION
    )

    bajas_qs, filtros = _filtrar_historial_construccion_qs(request, bajas_qs)

    if _usuario_es_cuadrilla(usuario):
        filtro = _filtro_cuadrilla_por_usuario(usuario)
        bajas_qs = bajas_qs.filter(filtro)

    bajas = []
    for baja in bajas_qs.order_by('-fecha'):
        observacion_campos = _extraer_campos_observacion(baja.observacion)
        coordenadas = observacion_campos.get('coordenadas', '')
        sector = observacion_campos.get('sector', '')
        identificadores_baja = [
            f"{item.identificador.tipo}: {item.identificador.valor}"
            for item in baja.identificadores_baja.select_related('identificador').all()
        ]
        if baja.serial_mac_retirado:
            identificadores_baja.append(f"RETIRADO: {baja.serial_mac_retirado}")

        bajas.append({
            'id': baja.id,
            'fecha': baja.fecha,
            'cuadrilla': baja.asignacion.cuadrilla.nombre,
            'usuario_registro': baja.usuario_registro or '',
            'producto': baja.producto_nombre or baja.asignacion.producto.nombre,
            'codigo': baja.producto_codigo or baja.asignacion.producto.codigo,
            'cliente': baja.cliente_nombre or (baja.cliente.nombre if baja.cliente else ''),
            'tipo_uso': baja.tipo_uso,
            'observacion': baja.observacion,
            'coordenadas': coordenadas,
            'sector': sector,
            'coord_punta_inicial': observacion_campos.get('coord_punta_inicial', ''),
            'coord_punta_final': observacion_campos.get('coord_punta_final', ''),
            'cantidad': baja.cantidad,
            'punta_inicial': baja.punta_inicial,
            'punta_final': baja.punta_final,
            'codigo_bobina': baja.codigo_bobina,
            'metraje': baja.metraje,
            'proveedor': baja.asignacion.producto.proveedor.nombre if baja.asignacion.producto.proveedor else '-',
            'detalle_mac': baja.detalle_mac,
            'detalle_serial': baja.detalle_serial,
            'serial_mac_retirado': baja.serial_mac_retirado,
            'estado_equipo': baja.estado_equipo,
            'es_equipo_activo': baja.es_equipo_activo,
            'retirado': baja.retirado,
            'fecha_retirado': baja.fecha_retirado,
            'identificadores': identificadores_baja,
        })

    return render(request, 'cuadrillas/historial_bajas.html', {
        'bajas': bajas,
        'titulo_historial': 'Historial de Bajas - Bodega Construccion',
        'limpiar_url_name': 'historial_bajas_construccion',
        'excel_url_name': 'reporte_historial_bajas_construccion_excel',
        'mostrar_excel': True,
        'mostrar_purgar': False,
        'mostrar_campos_construccion': True,
        'filtros_construccion': filtros,
    })


@role_required('admin', 'bodega')
def reporte_historial_bajas_construccion_excel(request):
    usuario = request.user
    bajas_qs = BajaMaterial.objects.select_related(
        'cliente',
        'asignacion__cuadrilla',
        'asignacion__producto',
        'asignacion__producto__proveedor',
    ).prefetch_related('identificadores_baja__identificador').filter(
        asignacion__producto__bodega=Producto.BODEGA_CONSTRUCCION
    )

    bajas_qs, _ = _filtrar_historial_construccion_qs(request, bajas_qs)

    if _usuario_es_cuadrilla(usuario):
        filtro = _filtro_cuadrilla_por_usuario(usuario)
        bajas_qs = bajas_qs.filter(filtro)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial Construccion'

    ws.append([
        'Fecha',
        'Sector',
        'Producto',
        'Codigo',
        'Arreglo/Construccion',
        'Coordenadas',
        'Coord punta inicial',
        'Coord punta final',
        'Cuadrilla',
        'Cantidad',
        'Estado equipo',
        'Punta inicial',
        'Punta final',
        'Codigo bobina',
        'Metraje total',
        'Observacion',
    ])

    for baja in bajas_qs.order_by('-fecha'):
        observacion_campos = _extraer_campos_observacion(baja.observacion)
        coordenadas = observacion_campos.get('coordenadas', '')
        sector = observacion_campos.get('sector', '')
        ws.append([
            timezone.localtime(baja.fecha).strftime('%d/%m/%Y %H:%M'),
            sector,
            baja.producto_nombre or baja.asignacion.producto.nombre,
            baja.producto_codigo or baja.asignacion.producto.codigo,
            baja.tipo_uso,
            coordenadas,
            observacion_campos.get('coord_punta_inicial', ''),
            observacion_campos.get('coord_punta_final', ''),
            baja.asignacion.cuadrilla.nombre,
            baja.cantidad,
            baja.estado_equipo or '',
            float(baja.punta_inicial or 0),
            float(baja.punta_final or 0),
            baja.codigo_bobina or '',
            float(baja.metraje or 0),
            baja.observacion or '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_bajas_construccion.xlsx"'
    wb.save(response)
    return response


@role_required('admin', 'bodega', 'soporte')
def equipos_retirados(request):
    proveedor_id = (request.GET.get('proveedor') or '').strip()
    producto_q = (request.GET.get('producto') or '').strip()
    serial_mac_q = _normalizar_texto(request.GET.get('serial_mac'))
    cliente_q = _normalizar_cliente_nombre(request.GET.get('cliente'))
    ct_q = _normalizar_texto(request.GET.get('ct'))
    texto_q = _normalizar_texto(request.GET.get('q'))
    origen_q = (request.GET.get('origen') or '').strip()
    fecha_desde = (request.GET.get('fecha_desde') or '').strip()
    fecha_hasta = (request.GET.get('fecha_hasta') or '').strip()
    proveedor_nombre_filtro = ''

    if request.method == 'POST':
        accion = (request.POST.get('accion') or '').strip()

        if accion == 'retirado':
            fecha_manual = _parse_fecha_retiro(request.POST.get('fecha'))

            producto_id = (request.POST.get('producto_id') or '').strip()
            proveedor_id_form = (request.POST.get('proveedor_id') or '').strip()
            cuadrilla_id = (request.POST.get('cuadrilla_id') or '').strip()
            cliente_id = (request.POST.get('cliente_id') or '').strip()
            categoria_id = (request.POST.get('categoria_id') or '').strip()
            marca_id = (request.POST.get('marca_id') or '').strip()
            modelo_id = (request.POST.get('modelo_id') or '').strip()

            producto = Producto.objects.select_related('proveedor', 'categoria', 'marca', 'modelo').filter(id=producto_id).first() if producto_id else None
            proveedor_obj = Proveedor.objects.filter(id=proveedor_id_form).first() if proveedor_id_form else None
            cuadrilla_obj = Cuadrilla.objects.filter(id=cuadrilla_id).first() if cuadrilla_id else None
            cliente_obj = Cliente.objects.filter(id=cliente_id).first() if cliente_id else None
            categoria_obj = Categoria.objects.filter(id=categoria_id).first() if categoria_id else None
            marca_obj = Marca.objects.filter(id=marca_id).first() if marca_id else None
            modelo_obj = ModeloEquipo.objects.filter(id=modelo_id).first() if modelo_id else None

            producto_nombre = _normalizar_texto(request.POST.get('producto_nombre')) or (producto.nombre if producto else '')
            producto_codigo = _normalizar_texto(request.POST.get('producto_codigo')) or (producto.codigo if producto else '')
            proveedor_nombre = _normalizar_texto(request.POST.get('proveedor_nombre'))
            if not proveedor_nombre:
                proveedor_nombre = (
                    proveedor_obj.nombre if proveedor_obj else (
                        producto.proveedor.nombre if producto and producto.proveedor else ''
                    )
                )
            cuadrilla_nombre = _normalizar_texto(request.POST.get('cuadrilla_nombre')) or (cuadrilla_obj.nombre if cuadrilla_obj else '')
            cliente_nombre = _normalizar_cliente_nombre(request.POST.get('cliente_nombre')) or (cliente_obj.nombre if cliente_obj else '')
            ct_cliente = _normalizar_texto(request.POST.get('ct_cliente')) or (cliente_obj.ct if cliente_obj else '')
            detalle_cambio = _normalizar_texto(request.POST.get('detalle_cambio'))
            serial_mac_retirado = _normalizar_texto(request.POST.get('serial_mac_retirado')).upper()
            codigo_retirado = _normalizar_texto(request.POST.get('codigo_retirado')).upper()
            categoria_retirado = _normalizar_texto(request.POST.get('categoria_retirado')) or (categoria_obj.nombre if categoria_obj else '')
            marca_retirado = _normalizar_texto(request.POST.get('marca_retirado')) or (marca_obj.nombre if marca_obj else '')
            modelo_retirado = _normalizar_texto(request.POST.get('modelo_retirado')) or (modelo_obj.nombre if modelo_obj else '')
            origen = 'Retirado'

            try:
                cantidad = int((request.POST.get('cantidad') or '1').strip() or '1')
            except ValueError:
                cantidad = 1
            metraje = Decimal('0')

            if not (producto_nombre or producto_codigo or serial_mac_retirado or codigo_retirado):
                messages.error(request, 'Debe ingresar al menos producto, codigo o serial/MAC para registrar el equipo retirado.')
                return redirect('equipos_retirados')

            if serial_mac_retirado and EquipoRetiradoManual.objects.filter(serial_mac_retirado__iexact=serial_mac_retirado).exists():
                messages.error(request, f'El serial/MAC {serial_mac_retirado} ya existe en equipos retirados.')
                return redirect('equipos_retirados')

            if not producto:
                producto = _buscar_producto_retiro(producto_codigo, producto_nombre)
            EquipoRetiradoManual.objects.create(
                fecha=fecha_manual,
                origen=origen,
                producto=producto,
                producto_nombre=producto_nombre or (producto.nombre if producto else ''),
                producto_codigo=producto_codigo or (producto.codigo if producto else ''),
                proveedor_nombre=proveedor_nombre or (producto.proveedor.nombre if producto and producto.proveedor else ''),
                cuadrilla_nombre=cuadrilla_nombre,
                cliente_nombre=cliente_nombre,
                ct_cliente=ct_cliente,
                detalle_cambio=detalle_cambio,
                categoria_retirado=categoria_retirado,
                marca_retirado=marca_retirado,
                modelo_retirado=modelo_retirado,
                usuario_registro=request.user.get_username(),
                serial_mac_retirado=serial_mac_retirado,
                codigo_retirado=codigo_retirado,
                cantidad=cantidad if cantidad > 0 else 1,
                metraje=metraje,
            )
            messages.success(request, 'Equipo retirado registrado correctamente.')
            return redirect('equipos_retirados')

    bajas_qs = BajaMaterial.objects.filter(
        tipo_uso='Cambio'
    ).exclude(
        serial_mac_retirado=''
    ).select_related(
        'cliente',
        'asignacion__cuadrilla',
        'asignacion__producto',
        'asignacion__producto__proveedor',
    )

    manuales_qs = EquipoRetiradoManual.objects.select_related('producto', 'producto__proveedor').all()

    if proveedor_id:
        proveedor_nombre_filtro = Proveedor.objects.filter(id=proveedor_id).values_list('nombre', flat=True).first() or ''
        bajas_qs = bajas_qs.filter(asignacion__producto__proveedor_id=proveedor_id)
        if proveedor_nombre_filtro:
            manuales_qs = manuales_qs.filter(
                Q(producto__proveedor_id=proveedor_id) |
                Q(proveedor_nombre__icontains=proveedor_nombre_filtro)
            )
        else:
            manuales_qs = manuales_qs.filter(producto__proveedor_id=proveedor_id)

    if producto_q:
        bajas_qs = bajas_qs.filter(
            Q(asignacion__producto__nombre__icontains=producto_q) |
            Q(asignacion__producto__codigo__icontains=producto_q)
        )
        manuales_qs = manuales_qs.filter(
            Q(producto_nombre__icontains=producto_q) |
            Q(producto_codigo__icontains=producto_q)
        )

    if cliente_q:
        bajas_qs = bajas_qs.filter(
            Q(cliente__nombre__iexact=cliente_q) |
            Q(cliente_nombre__iexact=cliente_q)
        )
        manuales_qs = manuales_qs.filter(cliente_nombre__icontains=cliente_q)

    if ct_q:
        bajas_qs = bajas_qs.filter(
            Q(cliente__ct__icontains=ct_q) |
            Q(cliente_nombre__icontains=ct_q)
        )
        manuales_qs = manuales_qs.filter(ct_cliente__icontains=ct_q)

    if serial_mac_q:
        bajas_qs = bajas_qs.filter(serial_mac_retirado__icontains=serial_mac_q)
        manuales_qs = manuales_qs.filter(serial_mac_retirado__icontains=serial_mac_q)

    if origen_q:
        if origen_q == 'Cambio':
            manuales_qs = manuales_qs.filter(origen='Cambio')
        elif origen_q == 'Retirado':
            bajas_qs = bajas_qs.none()
            # Compatibilidad con registros historicos que quedaron como "Manual".
            manuales_qs = manuales_qs.filter(origen__in=['Retirado', 'Manual'])
        else:
            bajas_qs = bajas_qs.none()
            manuales_qs = manuales_qs.none()

    if fecha_desde:
        fecha_inicio = _parse_fecha_retiro(fecha_desde)
        bajas_qs = bajas_qs.filter(fecha__date__gte=fecha_inicio.date())
        manuales_qs = manuales_qs.filter(fecha__date__gte=fecha_inicio.date())

    if fecha_hasta:
        fecha_fin = _parse_fecha_retiro(fecha_hasta)
        bajas_qs = bajas_qs.filter(fecha__date__lte=fecha_fin.date())
        manuales_qs = manuales_qs.filter(fecha__date__lte=fecha_fin.date())

    if texto_q:
        bajas_qs = bajas_qs.filter(
            Q(asignacion__producto__nombre__icontains=texto_q) |
            Q(asignacion__producto__codigo__icontains=texto_q) |
            Q(asignacion__cuadrilla__nombre__icontains=texto_q) |
            Q(cliente_nombre__icontains=texto_q) |
            Q(cliente__nombre__icontains=texto_q) |
            Q(asignacion__producto__proveedor__nombre__icontains=texto_q) |
            Q(codigo_bobina__icontains=texto_q) |
            Q(serial_mac_retirado__icontains=texto_q) |
            Q(categoria_retirado__icontains=texto_q) |
            Q(marca_retirado__icontains=texto_q) |
            Q(modelo_retirado__icontains=texto_q) |
            Q(observacion__icontains=texto_q)
        )
        manuales_qs = manuales_qs.filter(
            Q(producto_nombre__icontains=texto_q) |
            Q(producto_codigo__icontains=texto_q) |
            Q(proveedor_nombre__icontains=texto_q) |
            Q(cuadrilla_nombre__icontains=texto_q) |
            Q(cliente_nombre__icontains=texto_q) |
            Q(ct_cliente__icontains=texto_q) |
            Q(detalle_cambio__icontains=texto_q) |
            Q(serial_mac_retirado__icontains=texto_q) |
            Q(codigo_retirado__icontains=texto_q) |
            Q(categoria_retirado__icontains=texto_q) |
            Q(marca_retirado__icontains=texto_q) |
            Q(modelo_retirado__icontains=texto_q)
        )

    registros = []
    for baja in bajas_qs.order_by('-fecha'):
        tipo_id = _inferir_tipo_identificador(baja.serial_mac_retirado)
        disponible = DetalleIdentificador.objects.filter(
            producto=baja.asignacion.producto,
            tipo=tipo_id,
            valor__iexact=baja.serial_mac_retirado,
            asignacion_detalle__isnull=True,
        ).exists()
        registros.append({
            'id': baja.id,
            'fecha': baja.fecha,
            'origen': 'Cambio',
            'producto': baja.asignacion.producto.nombre,
            'codigo': baja.asignacion.producto.codigo,
            'proveedor': baja.asignacion.producto.proveedor.nombre if baja.asignacion.producto.proveedor else '-',
            'cuadrilla': baja.asignacion.cuadrilla.nombre,
            'usuario_registro': baja.usuario_registro or '-',
            'cliente': baja.cliente_nombre or (baja.cliente.nombre if baja.cliente else '-'),
            'ct': baja.cliente.ct if baja.cliente else '',
            'detalle_cambio': baja.observacion or '-',
            'tipo_identificador': tipo_id,
            'serial_mac_retirado': baja.serial_mac_retirado,
            'codigo_retirado': baja.codigo_retirado,
            'categoria_retirado': baja.categoria_retirado,
            'marca_retirado': baja.marca_retirado,
            'modelo_retirado': baja.modelo_retirado,
            'reutilizado': baja.reutilizado,
            'fecha_reutilizado': baja.fecha_reutilizado,
            'estado_bodega': 'Disponible' if disponible else 'Asignado',
        })

    for retiro in manuales_qs.order_by('-fecha'):
        tiene_datos_para_reuso = bool(
            (retiro.serial_mac_retirado or '').strip() and
            (retiro.producto_id or (retiro.producto_codigo or '').strip() or (retiro.producto_nombre or '').strip())
        )
        origen_retiro = 'Retirado' if (retiro.origen or '').strip().lower() == 'manual' else retiro.origen
        registros.append({
            'id': retiro.id,
            'fecha': retiro.fecha,
            'origen': origen_retiro,
            'producto': retiro.producto_nombre or (retiro.producto.nombre if retiro.producto else '-'),
            'codigo': retiro.producto_codigo or (retiro.producto.codigo if retiro.producto else '-'),
            'proveedor': retiro.proveedor_nombre or (retiro.producto.proveedor.nombre if retiro.producto and retiro.producto.proveedor else '-'),
            'cuadrilla': retiro.cuadrilla_nombre or '-',
            'usuario_registro': retiro.usuario_registro or '-',
            'cliente': retiro.cliente_nombre or '-',
            'ct': retiro.ct_cliente or '',
            'detalle_cambio': retiro.detalle_cambio or '-',
            'tipo_identificador': _inferir_tipo_identificador(retiro.serial_mac_retirado) if retiro.serial_mac_retirado else '-',
            'serial_mac_retirado': retiro.serial_mac_retirado or '-',
            'codigo_retirado': retiro.codigo_retirado or '-',
            'categoria_retirado': retiro.categoria_retirado or '-',
            'marca_retirado': retiro.marca_retirado or '-',
            'modelo_retirado': retiro.modelo_retirado or '-',
            'reutilizado': retiro.reutilizado,
            'fecha_reutilizado': retiro.fecha_reutilizado,
            'estado_bodega': 'Retirado',
            'es_manual': True,
            'puede_reutilizar': tiene_datos_para_reuso,
        })

    registros.sort(key=lambda item: item['fecha'], reverse=True)

    if request.GET.get('export') == 'excel':
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Equipos retirados'
        headers = [
            'Fecha', 'Origen', 'Producto', 'Codigo', 'Proveedor', 'Cuadrilla',
            'Usuario registro', 'Cliente', 'CT', 'Detalle de cambio', 'Categoria retirado',
            'Marca', 'Modelo', 'Serial/MAC retirado', 'Codigo retirado', 'Estado'
        ]
        sheet.append(headers)

        for r in registros:
            fecha_texto = timezone.localtime(r['fecha']).strftime('%d/%m/%Y %H:%M') if r.get('fecha') else ''
            sheet.append([
                fecha_texto,
                r.get('origen', ''),
                r.get('producto', ''),
                r.get('codigo', ''),
                r.get('proveedor', ''),
                r.get('cuadrilla', ''),
                r.get('usuario_registro', ''),
                r.get('cliente', ''),
                r.get('ct', ''),
                r.get('detalle_cambio', ''),
                r.get('categoria_retirado', ''),
                r.get('marca_retirado', ''),
                r.get('modelo_retirado', ''),
                r.get('serial_mac_retirado', ''),
                r.get('codigo_retirado', ''),
                'Reutilizado' if r.get('reutilizado') else r.get('estado_bodega', ''),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="equipos_retirados.xlsx"'
        workbook.save(response)
        return response

    proveedores = Proveedor.objects.order_by('nombre')
    productos_busqueda = Producto.objects.select_related('proveedor').order_by('nombre')
    clientes_busqueda = Cliente.objects.order_by('nombre')
    cuadrillas_busqueda = Cuadrilla.objects.order_by('nombre')
    categorias_busqueda = Categoria.objects.order_by('nombre')
    marcas_busqueda = Marca.objects.order_by('nombre')
    modelos_busqueda = ModeloEquipo.objects.order_by('nombre')

    return render(request, 'cuadrillas/equipos_retirados.html', {
        'registros': registros,
        'proveedores': proveedores,
        'productos_busqueda': productos_busqueda,
        'clientes_busqueda': clientes_busqueda,
        'cuadrillas_busqueda': cuadrillas_busqueda,
        'categorias_busqueda': categorias_busqueda,
        'marcas_busqueda': marcas_busqueda,
        'modelos_busqueda': modelos_busqueda,
        'proveedor_seleccionado': proveedor_id,
        'producto_q': producto_q,
        'cliente_q': cliente_q,
        'ct_q': ct_q,
        'serial_mac_q': serial_mac_q,
        'texto_q': texto_q,
        'origen_q': origen_q,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    })


@role_required('admin', 'bodega')
def reutilizar_equipo_retirado(request, baja_id):
    if request.method != 'POST':
        return redirect('equipos_retirados')

    baja = get_object_or_404(
        BajaMaterial.objects.select_related('asignacion__producto'),
        id=baja_id,
        tipo_uso='Cambio',
    )

    if not baja.serial_mac_retirado:
        messages.error(request, 'Esta baja no tiene equipo retirado registrado.')
        return redirect('equipos_retirados')

    if baja.reutilizado:
        messages.info(request, 'Este equipo ya fue enviado a bodega principal anteriormente.')
        return redirect('equipos_retirados')

    producto = baja.asignacion.producto
    tipo_retirado = _inferir_tipo_identificador(baja.serial_mac_retirado)
    codigo_retirado = (baja.codigo_retirado or '').strip() or None
    detalle, creado = DetalleIdentificador.objects.get_or_create(
        producto=producto,
        tipo=tipo_retirado,
        valor=baja.serial_mac_retirado,
        defaults={
            'codigo_individual': codigo_retirado,
            'metraje': Decimal('0'),
        }
    )

    if not creado and codigo_retirado and not (detalle.codigo_individual or '').strip():
        detalle.codigo_individual = codigo_retirado
        detalle.save(update_fields=['codigo_individual'])

    if creado:
        producto.stock = (producto.stock or 0) + 1
        producto.save(update_fields=['stock'])

    baja.reutilizado = True
    baja.fecha_reutilizado = timezone.now()
    baja.save(update_fields=['reutilizado', 'fecha_reutilizado'])

    messages.success(request, f'Equipo {baja.serial_mac_retirado} enviado a bodega principal correctamente.')
    return redirect('equipos_retirados')


@role_required('admin', 'bodega')
def reutilizar_equipo_retirado_manual(request, retiro_id):
    if request.method != 'POST':
        return redirect('equipos_retirados')

    retiro = get_object_or_404(EquipoRetiradoManual.objects.select_related('producto'), id=retiro_id)

    if retiro.reutilizado:
        messages.info(request, 'Este equipo retirado ya fue enviado a bodega principal anteriormente.')
        return redirect('equipos_retirados')

    serial_mac = (retiro.serial_mac_retirado or '').strip().upper()
    if not serial_mac:
        messages.error(request, 'El registro retirado no tiene serial/MAC para poder reutilizarlo.')
        return redirect('equipos_retirados')

    producto = retiro.producto
    if not producto:
        producto = _buscar_producto_retiro(retiro.producto_codigo, retiro.producto_nombre)
        if producto:
            retiro.producto = producto

    if not producto:
        messages.error(request, 'No se pudo identificar el producto para este equipo retirado. Complete producto/codigo y vuelva a intentar.')
        return redirect('equipos_retirados')

    tipo_retirado = _inferir_tipo_identificador(serial_mac)
    codigo_retirado = (retiro.codigo_retirado or '').strip() or None
    detalle, creado = DetalleIdentificador.objects.get_or_create(
        producto=producto,
        tipo=tipo_retirado,
        valor=serial_mac,
        defaults={
            'codigo_individual': codigo_retirado,
            'metraje': Decimal('0'),
        }
    )

    if not creado and codigo_retirado and not (detalle.codigo_individual or '').strip():
        detalle.codigo_individual = codigo_retirado
        detalle.save(update_fields=['codigo_individual'])

    if creado:
        producto.stock = (producto.stock or 0) + 1
        producto.save(update_fields=['stock'])

    retiro.reutilizado = True
    retiro.fecha_reutilizado = timezone.now()
    retiro.save(update_fields=['producto', 'reutilizado', 'fecha_reutilizado'])

    messages.success(request, f'Equipo retirado {serial_mac} enviado a bodega principal correctamente.')
    return redirect('equipos_retirados')


@role_required('admin', 'bodega')
def retirar_cliente_baja(request, baja_id):
    if request.method != 'POST':
        return redirect('historial_bajas')

    baja = BajaMaterial.objects.select_related('cliente', 'asignacion__producto').filter(id=baja_id).first()
    if not baja:
        messages.error(request, 'El registro de baja no existe o ya no esta disponible.')
        return redirect('historial_bajas')

    if not baja.es_equipo_activo:
        messages.error(request, 'La accion Retirado solo aplica para productos activos.')
        return redirect('historial_bajas')

    if baja.retirado:
        messages.info(request, 'Esta baja ya fue marcada como Retirado.')
        return redirect('historial_bajas')

    producto = baja.asignacion.producto
    producto.stock = (producto.stock or 0) + (baja.cantidad or 0)
    producto.metraje = (producto.metraje or Decimal('0')) + (baja.metraje or Decimal('0'))
    producto.save()

    BajaIdentificador.objects.filter(baja=baja).delete()

    cliente = baja.cliente
    nombre_cliente = _normalizar_cliente_nombre(
        baja.cliente_nombre or (cliente.nombre if cliente else '')
    )

    if cliente:
        cliente.delete()

    if nombre_cliente:
        Cliente.objects.filter(nombre__iexact=nombre_cliente).delete()

    baja.retirado = True
    baja.fecha_retirado = timezone.now()
    baja.cliente = None
    baja.save(update_fields=['retirado', 'fecha_retirado', 'cliente'])

    messages.success(request, 'Registro marcado como Retirado. Cliente eliminado y equipo devuelto a bodega correctamente.')
    return redirect('historial_bajas')


@role_required('admin')
def purgar_historial_bajas(request):
    if request.method != 'POST':
        return redirect('historial_bajas')

    fecha_corte = timezone.now() - timedelta(days=183)
    bajas_antiguas = BajaMaterial.objects.filter(fecha__lt=fecha_corte)
    total_bajas = bajas_antiguas.count()

    if total_bajas == 0:
        messages.info(request, 'No existen registros de bajas con más de 6 meses para purgar.')
        return redirect('historial_bajas')

    eliminados, _ = bajas_antiguas.delete()
    messages.success(
        request,
        f'Purga completada. Se eliminaron {total_bajas} baja(s) con más de 6 meses. Registros afectados en cascada: {eliminados}.'
    )
    return redirect('historial_bajas')


@role_required('admin', 'bodega')
def reporte_historial_bajas_excel(request):
    usuario = request.user
    cliente_nombre = _normalizar_cliente_nombre(request.GET.get('cliente'))
    proveedor_id = request.GET.get('proveedor', '').strip()
    es_construccion = (request.GET.get('construccion') or '').strip() == '1'

    bajas_qs = BajaMaterial.objects.select_related(
        'cliente',
        'asignacion__cuadrilla',
        'asignacion__producto',
        'asignacion__producto__proveedor',
    ).prefetch_related(
        'identificadores_baja__identificador'
    )

    if es_construccion:
        bajas_qs = bajas_qs.filter(asignacion__producto__bodega=Producto.BODEGA_CONSTRUCCION)
        bajas_qs, _ = _filtrar_historial_construccion_qs(request, bajas_qs)
    else:
        bajas_qs = bajas_qs.exclude(asignacion__producto__bodega=Producto.BODEGA_CONSTRUCCION)

    if _usuario_es_cuadrilla(usuario):
        filtro = _filtro_cuadrilla_por_usuario(usuario)
        bajas_qs = bajas_qs.filter(filtro)

    if (not es_construccion) and cliente_nombre:
        bajas_qs = bajas_qs.filter(cliente__nombre__iexact=cliente_nombre)

    if (not es_construccion) and proveedor_id:
        bajas_qs = bajas_qs.filter(asignacion__producto__proveedor_id=proveedor_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial Bajas'

    if es_construccion:
        ws.append([
            'Fecha',
            'Sector',
            'Producto',
            'Codigo',
            'Arreglo/Construccion',
            'Coordenadas',
            'Coord punta inicial',
            'Coord punta final',
            'Cuadrilla',
            'Cantidad',
            'Estado equipo',
            'Punta inicial',
            'Punta final',
            'Codigo bobina',
            'Metraje total',
            'Observacion',
        ])
    else:
        ws.append([
            'Fecha',
            'Cliente',
            'Tipo de uso',
            'Detalle cambio',
            'Cuadrilla',
            'Producto',
            'Proveedor',
            'Codigo',
            'Cantidad',
            'Estado equipo',
            'Punta inicial',
            'Punta final',
            'Coord punta inicial',
            'Coord punta final',
            'Coord cliente',
            'Coord caja',
            'Codigo bobina',
            'Metraje',
            'MAC',
            'Serial',
            'Serial/MAC retirado',
            'Seriales/MAC',
        ])

    for baja in bajas_qs.order_by('-fecha'):
        identificadores = ', '.join([
            f"{item.identificador.tipo}: {item.identificador.valor}"
            for item in baja.identificadores_baja.select_related('identificador').all()
        ])

        if es_construccion:
            observacion_campos = _extraer_campos_observacion(baja.observacion)
            coordenadas = observacion_campos.get('coordenadas', '')
            sector = observacion_campos.get('sector', '')
            ws.append([
                timezone.localtime(baja.fecha).strftime('%d/%m/%Y %H:%M'),
                sector,
                baja.producto_nombre or baja.asignacion.producto.nombre,
                baja.producto_codigo or baja.asignacion.producto.codigo,
                baja.tipo_uso,
                coordenadas,
                observacion_campos.get('coord_punta_inicial', ''),
                observacion_campos.get('coord_punta_final', ''),
                baja.asignacion.cuadrilla.nombre,
                baja.cantidad,
                baja.estado_equipo or '',
                float(baja.punta_inicial or 0),
                float(baja.punta_final or 0),
                baja.codigo_bobina or '',
                float(baja.metraje or 0),
                baja.observacion or '',
            ])
        else:
            observacion_campos = _extraer_campos_observacion(baja.observacion)
            ws.append([
                timezone.localtime(baja.fecha).strftime('%d/%m/%Y %H:%M'),
                baja.cliente_nombre or (baja.cliente.nombre if baja.cliente else ''),
                baja.tipo_uso,
                _detalle_observacion_sin_coordenadas(baja.observacion),
                baja.asignacion.cuadrilla.nombre,
                baja.producto_nombre or baja.asignacion.producto.nombre,
                baja.asignacion.producto.proveedor.nombre if baja.asignacion.producto.proveedor else '',
                baja.producto_codigo or baja.asignacion.producto.codigo,
                baja.cantidad,
                baja.estado_equipo or '',
                float(baja.punta_inicial or 0),
                float(baja.punta_final or 0),
                observacion_campos.get('coord_punta_inicial', ''),
                observacion_campos.get('coord_punta_final', ''),
                observacion_campos.get('coord_cliente', ''),
                observacion_campos.get('coord_caja', ''),
                baja.codigo_bobina or '',
                float(baja.metraje or 0),
                baja.detalle_mac,
                baja.detalle_serial,
                baja.serial_mac_retirado,
                identificadores,
            ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_bajas.xlsx"'
    wb.save(response)
    return response

@login_required
def mis_asignaciones(request):
    usuario = request.user

    # Obtener la cuadrilla del usuario (grupo)
    grupo = usuario.groups.first()

    asignaciones = AsignacionMaterial.objects.none()

    if grupo:
        asignaciones = AsignacionMaterial.objects.filter(
            cuadrilla__nombre=grupo.name
        )

    return render(request, 'cuadrillas/mis_asignaciones.html', {
        'asignaciones': asignaciones
    })