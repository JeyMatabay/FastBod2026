from django.shortcuts import render, redirect, get_object_or_404
from decimal import Decimal, InvalidOperation
from .models import Producto, DetalleIdentificador, Proveedor
from .forms import (
    ProductoForm,
    DetalleIdentificadorForm,
    ProveedorForm,
    es_categoria_activa,
    es_categoria_fibra,
    es_categoria_con_metraje,
    es_categoria_codigo_unico,
)
from cuadrillas.models import Cuadrilla, AsignacionMaterial
from cuadrillas.models import AsignacionIdentificador, BajaIdentificador
import csv
from pathlib import Path
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from django.db.models import F
from django.db import connection
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db import IntegrityError
from django.conf import settings
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from django.utils import timezone
from pathlib import Path as FilePath
from usuarios.decorators import role_required
import os
import psutil

DJANGO_PROCESS = psutil.Process(os.getpid())
DJANGO_PROCESS.cpu_percent(None)


def _normalizar_identificador(valor):
    return (str(valor or '').strip().upper())


def _parse_metraje(valor):
    texto = str(valor or '').strip().replace(',', '.')
    if not texto:
        return Decimal('0')
    try:
        numero = Decimal(texto)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')
    return numero if numero >= 0 else Decimal('0')


def _format_bytes(valor):
    size = float(valor or 0)
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    for unit in units:
        if size < 1024 or unit == units[-1]:
            if unit == 'B':
                return f"{int(size)} {unit}"
            return f"{size:.2f} {unit}"
        size /= 1024


def _postgres_snapshot():
    if connection.vendor != 'postgresql':
        return None

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    pg_database_size(current_database()) AS database_size,
                    (SELECT COUNT(*) FROM pg_stat_activity WHERE datname = current_database()) AS total_connections,
                    (SELECT COUNT(*) FROM pg_stat_activity WHERE datname = current_database() AND state = 'active') AS active_connections,
                    (SELECT COUNT(*) FROM pg_stat_activity WHERE datname = current_database() AND wait_event IS NOT NULL) AS waiting_connections,
                    xact_commit,
                    xact_rollback,
                    tup_returned,
                    tup_fetched,
                    tup_inserted,
                    tup_updated,
                    tup_deleted,
                    blks_read,
                    blks_hit,
                    temp_bytes,
                    deadlocks
                FROM pg_stat_database
                WHERE datname = current_database()
                """
            )
            row = cursor.fetchone()

        if not row:
            return None

        total_blocks = (row[11] or 0) + (row[12] or 0)
        cache_hit_ratio = 0
        if total_blocks > 0:
            cache_hit_ratio = round(((row[12] or 0) / total_blocks) * 100, 1)

        return {
            'database_size': _format_bytes(row[0]),
            'total_connections': row[1] or 0,
            'active_connections': row[2] or 0,
            'waiting_connections': row[3] or 0,
            'xact_commit': row[4] or 0,
            'xact_rollback': row[5] or 0,
            'tup_returned': row[6] or 0,
            'tup_fetched': row[7] or 0,
            'tup_inserted': row[8] or 0,
            'tup_updated': row[9] or 0,
            'tup_deleted': row[10] or 0,
            'blks_read': row[11] or 0,
            'blks_hit': row[12] or 0,
            'temp_bytes': _format_bytes(row[13] or 0),
            'deadlocks': row[14] or 0,
            'cache_hit_ratio': cache_hit_ratio,
        }
    except Exception:
        return None


def _resource_snapshot():
    memoria_virtual = psutil.virtual_memory()
    disco = psutil.disk_usage(str(settings.BASE_DIR))
    memoria_proceso = DJANGO_PROCESS.memory_info()
    ahora = timezone.now()
    inicio_proceso = timezone.datetime.fromtimestamp(DJANGO_PROCESS.create_time(), tz=timezone.get_current_timezone())
    inicio_sistema = timezone.datetime.fromtimestamp(psutil.boot_time(), tz=timezone.get_current_timezone())

    try:
        archivos_abiertos = len(DJANGO_PROCESS.open_files())
    except (psutil.AccessDenied, psutil.NoSuchProcess):
        archivos_abiertos = None

    return {
        'timestamp': ahora.isoformat(),
        'timestamp_texto': ahora.strftime('%d/%m/%Y %H:%M:%S'),
        'cpu_percent': round(psutil.cpu_percent(interval=None), 1),
        'cpu_count_logical': psutil.cpu_count() or 0,
        'cpu_count_physical': psutil.cpu_count(logical=False) or 0,
        'process_cpu_percent': round(DJANGO_PROCESS.cpu_percent(interval=None), 1),
        'memory_percent': round(memoria_virtual.percent, 1),
        'memory_used': _format_bytes(memoria_virtual.used),
        'memory_total': _format_bytes(memoria_virtual.total),
        'disk_percent': round(disco.percent, 1),
        'disk_used': _format_bytes(disco.used),
        'disk_total': _format_bytes(disco.total),
        'process_memory': _format_bytes(memoria_proceso.rss),
        'process_threads': DJANGO_PROCESS.num_threads(),
        'open_files': archivos_abiertos,
        'system_uptime': str(ahora - inicio_sistema).split('.')[0],
        'process_uptime': str(ahora - inicio_proceso).split('.')[0],
        'postgres': _postgres_snapshot(),
    }


def _leer_identificadores_archivo(archivo):
    extension = Path(archivo.name).suffix.lower()
    filas = []

    if extension == '.csv':
        contenido = archivo.read().decode('utf-8-sig', errors='ignore').splitlines()
        reader = csv.reader(contenido)
        for row in reader:
            if not row:
                continue
            codigo = _normalizar_identificador(row[0]) if len(row) > 0 else ''
            valor = _normalizar_identificador(row[1]) if len(row) > 1 else ''
            if codigo or valor:
                filas.append({'codigo': codigo, 'valor': valor})
        return filas

    if extension == '.xlsx':
        from openpyxl import load_workbook

        workbook = load_workbook(archivo, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            codigo = _normalizar_identificador(row[0]) if len(row) > 0 else ''
            valor = _normalizar_identificador(row[1]) if len(row) > 1 else ''
            if codigo or valor:
                filas.append({'codigo': codigo, 'valor': valor})
        return filas

    raise ValueError('Formato no soportado. Use .csv o .xlsx')


@role_required('admin', 'bodega')
def lista_productos(request):

    modelo_filtro = (request.GET.get('modelo') or '').strip()
    productos = Producto.objects.filter(bodega=Producto.BODEGA_GENERAL)
    if modelo_filtro == '__sin_modelo__':
        productos = productos.filter(modelo__isnull=True)
    elif modelo_filtro:
        productos = productos.filter(modelo__nombre__iexact=modelo_filtro)

    productos_stock_bajo = Producto.objects.filter(
        bodega=Producto.BODEGA_GENERAL,
        stock__lte=F('stock_minimo')
    ).order_by('nombre')
    totales_por_modelo = Producto.objects.filter(bodega=Producto.BODEGA_GENERAL).values('modelo_id', 'modelo__nombre').annotate(
        total=Coalesce(Sum('stock'), 0)
    ).order_by('modelo__nombre')

    return render(request, 'inventario/productos.html', {
        'productos': productos,
        'totales_por_modelo': totales_por_modelo,
        'productos_stock_bajo': productos_stock_bajo,
        'total_stock_bajo': productos_stock_bajo.count(),
        'modelo_filtro': modelo_filtro,
        'titulo_inventario': 'Inventario de Productos',
        'crear_url': '/crear_producto/',
        'detalle_prefix': '/producto/',
        'editar_prefix': '/editar_producto/',
        'eliminar_prefix': '/eliminar_producto/',
    })

@role_required('admin', 'bodega')
def crear_producto(request):

    form = ProductoForm(request.POST or None)

    if form.is_valid():

        nuevo_producto = form.save(commit=False)
        nuevo_producto.bodega = Producto.BODEGA_GENERAL
        codigo_normalizado = (nuevo_producto.codigo or '').strip()
        nombre_normalizado = (nuevo_producto.nombre or '').strip()
        es_activo = es_categoria_activa(nuevo_producto.categoria.nombre if nuevo_producto.categoria else '')
        es_metraje = es_categoria_con_metraje(nuevo_producto.categoria.nombre if nuevo_producto.categoria else '')

        existente = None
        if (not es_activo) and nombre_normalizado and nuevo_producto.categoria_id:
            existente = Producto.objects.filter(
                nombre__iexact=nombre_normalizado,
                categoria=nuevo_producto.categoria,
                bodega=nuevo_producto.bodega,
            ).first()

        if existente:
            existente.stock += (nuevo_producto.stock or 0)
            if es_metraje:
                existente.metraje += (nuevo_producto.metraje or 0)

            # Completa datos faltantes del producto existente sin crear duplicado.
            if not existente.mac and nuevo_producto.mac:
                existente.mac = nuevo_producto.mac
            if not existente.serial and nuevo_producto.serial:
                existente.serial = nuevo_producto.serial
            if not existente.marca and nuevo_producto.marca:
                existente.marca = nuevo_producto.marca
            if not existente.modelo and nuevo_producto.modelo:
                existente.modelo = nuevo_producto.modelo
            if not existente.tipo_fibra and nuevo_producto.tipo_fibra:
                existente.tipo_fibra = nuevo_producto.tipo_fibra
            if not existente.hilo and nuevo_producto.hilo:
                existente.hilo = nuevo_producto.hilo
            if not existente.codigo and codigo_normalizado:
                existente.codigo = codigo_normalizado

            existente.save()
            messages.success(request, 'Producto existente: se sumo la cantidad al stock actual.')
        else:
            nuevo_producto.codigo = codigo_normalizado
            nuevo_producto.nombre = nombre_normalizado
            nuevo_producto.save()
            messages.success(request, 'Producto creado correctamente.')

        return redirect('/productos/')

    return render(request, 'inventario/crear_producto.html', {
        'form': form,
        'titulo_form': 'Agregar Producto',
        'cancelar_url': '/productos/',
    })


@role_required('admin', 'bodega')
def editar_producto(request, id):

    producto = get_object_or_404(Producto, id=id, bodega=Producto.BODEGA_GENERAL)

    form = ProductoForm(request.POST or None, instance=producto)

    if form.is_valid():

        form.save()

        return redirect('/productos/')

    return render(request, 'inventario/editar_producto.html', {
        'form': form,
        'titulo_form': 'Editar Producto',
        'cancelar_url': '/productos/',
    })


@role_required('admin', 'bodega')
def eliminar_producto(request, id):

    producto = get_object_or_404(Producto, id=id, bodega=Producto.BODEGA_GENERAL)

    producto.delete()

    return redirect('/productos/')


@role_required('admin', 'bodega')
def lista_productos_construccion(request):

    modelo_filtro = (request.GET.get('modelo') or '').strip()
    productos = Producto.objects.filter(bodega=Producto.BODEGA_CONSTRUCCION)
    if modelo_filtro == '__sin_modelo__':
        productos = productos.filter(modelo__isnull=True)
    elif modelo_filtro:
        productos = productos.filter(modelo__nombre__iexact=modelo_filtro)

    productos_stock_bajo = Producto.objects.filter(
        bodega=Producto.BODEGA_CONSTRUCCION,
        stock__lte=F('stock_minimo')
    ).order_by('nombre')
    totales_por_modelo = Producto.objects.filter(bodega=Producto.BODEGA_CONSTRUCCION).values('modelo_id', 'modelo__nombre').annotate(
        total=Coalesce(Sum('stock'), 0)
    ).order_by('modelo__nombre')

    return render(request, 'inventario/productos.html', {
        'productos': productos,
        'totales_por_modelo': totales_por_modelo,
        'productos_stock_bajo': productos_stock_bajo,
        'total_stock_bajo': productos_stock_bajo.count(),
        'modelo_filtro': modelo_filtro,
        'titulo_inventario': 'Bodega de Equipos de Construccion',
        'crear_url': '/crear_producto_construccion/',
        'detalle_prefix': '/producto-construccion/',
        'editar_prefix': '/editar_producto_construccion/',
        'eliminar_prefix': '/eliminar_producto_construccion/',
    })


@role_required('admin', 'bodega')
def crear_producto_construccion(request):

    form = ProductoForm(request.POST or None)

    if form.is_valid():

        nuevo_producto = form.save(commit=False)
        nuevo_producto.bodega = Producto.BODEGA_CONSTRUCCION
        codigo_normalizado = (nuevo_producto.codigo or '').strip()
        nombre_normalizado = (nuevo_producto.nombre or '').strip()
        es_activo = es_categoria_activa(nuevo_producto.categoria.nombre if nuevo_producto.categoria else '')
        es_metraje = es_categoria_con_metraje(nuevo_producto.categoria.nombre if nuevo_producto.categoria else '')

        existente = None
        if (not es_activo) and nombre_normalizado and nuevo_producto.categoria_id:
            existente = Producto.objects.filter(
                nombre__iexact=nombre_normalizado,
                categoria=nuevo_producto.categoria,
                bodega=nuevo_producto.bodega,
            ).first()

        if existente:
            existente.stock += (nuevo_producto.stock or 0)
            if es_metraje:
                existente.metraje += (nuevo_producto.metraje or 0)

            if not existente.mac and nuevo_producto.mac:
                existente.mac = nuevo_producto.mac
            if not existente.serial and nuevo_producto.serial:
                existente.serial = nuevo_producto.serial
            if not existente.marca and nuevo_producto.marca:
                existente.marca = nuevo_producto.marca
            if not existente.modelo and nuevo_producto.modelo:
                existente.modelo = nuevo_producto.modelo
            if not existente.tipo_fibra and nuevo_producto.tipo_fibra:
                existente.tipo_fibra = nuevo_producto.tipo_fibra
            if not existente.hilo and nuevo_producto.hilo:
                existente.hilo = nuevo_producto.hilo
            if not existente.codigo and codigo_normalizado:
                existente.codigo = codigo_normalizado

            existente.save()
            messages.success(request, 'Producto de construccion existente: se sumo la cantidad al stock actual.')
        else:
            nuevo_producto.codigo = codigo_normalizado
            nuevo_producto.nombre = nombre_normalizado
            nuevo_producto.save()
            messages.success(request, 'Producto de construccion creado correctamente.')

        return redirect('/productos-construccion/')

    return render(request, 'inventario/crear_producto.html', {
        'form': form,
        'titulo_form': 'Agregar Producto - Bodega Construccion',
        'cancelar_url': '/productos-construccion/',
    })


@role_required('admin', 'bodega')
def editar_producto_construccion(request, id):

    producto = get_object_or_404(Producto, id=id, bodega=Producto.BODEGA_CONSTRUCCION)

    form = ProductoForm(request.POST or None, instance=producto)

    if form.is_valid():
        form.save()
        return redirect('/productos-construccion/')

    return render(request, 'inventario/editar_producto.html', {
        'form': form,
        'titulo_form': 'Editar Producto - Bodega Construccion',
        'cancelar_url': '/productos-construccion/',
    })


@role_required('admin', 'bodega')
def eliminar_producto_construccion(request, id):

    producto = get_object_or_404(Producto, id=id, bodega=Producto.BODEGA_CONSTRUCCION)
    producto.delete()

    return redirect('/productos-construccion/')


@role_required('admin', 'bodega', 'cuadrilla')
def detalle_producto(request, id):

    producto = get_object_or_404(Producto, id=id)
    detalle_url = f'/producto-construccion/{producto.id}/' if producto.bodega == Producto.BODEGA_CONSTRUCCION else f'/producto/{producto.id}/'
    filtro_estado = (request.GET.get('estado') or '').strip().lower()
    estados_validos = {'bodega', 'utilizado', 'cuadrilla'}
    if filtro_estado not in estados_validos:
        filtro_estado = ''
    es_activo = es_categoria_activa(producto.categoria.nombre if producto.categoria else '')
    es_fibra = es_categoria_fibra(producto.categoria.nombre if producto.categoria else '')
    es_metraje = es_categoria_con_metraje(producto.categoria.nombre if producto.categoria else '')
    es_cable_metraje = es_metraje and not es_activo
    es_codigo_unico = es_categoria_codigo_unico(producto.categoria.nombre if producto.categoria else '')

    if request.method == 'POST' and (es_activo or es_cable_metraje or es_codigo_unico):
        accion = request.POST.get('accion')

        if accion == 'masivo':
            archivo = request.FILES.get('archivo_identificadores')
            tipo_masivo = _normalizar_identificador(request.POST.get('tipo_masivo'))

            if not archivo:
                messages.error(request, 'Debe seleccionar un archivo CSV o XLSX.')
            elif es_activo and tipo_masivo not in ['SERIAL', 'MAC']:
                messages.error(request, 'Debe seleccionar el tipo de identificador para carga masiva.')
            else:
                try:
                    filas = _leer_identificadores_archivo(archivo)
                except Exception as exc:
                    messages.error(request, f'No se pudo leer el archivo: {exc}')
                else:
                    headers_codigo = {'CODIGO', 'CÓDIGO', 'COD', 'PRODUCTO', 'SKU'}
                    headers_valor = {'SERIAL', 'MAC', 'LOTE', 'LOTE/N DE SERIE', 'VALOR', 'IDENTIFICADOR', 'METRAJE', 'METROS'}

                    limpios = []
                    for fila in filas:
                        codigo_fila = fila.get('codigo', '')
                        valor_fila = fila.get('valor', '')

                        if not codigo_fila and not valor_fila:
                            continue

                        if codigo_fila in headers_codigo or valor_fila in headers_valor:
                            continue

                        limpios.append({'codigo': codigo_fila, 'valor': valor_fila})

                    vistos = set()
                    unicos = []
                    for item in limpios:
                        llave = (item['codigo'], item['valor'])
                        if llave not in vistos:
                            vistos.add(llave)
                            unicos.append(item)

                    cantidad_registrada = DetalleIdentificador.objects.filter(producto=producto).count()
                    capacidad_disponible = max(producto.stock - cantidad_registrada, 0)
                    metraje_registrado = DetalleIdentificador.objects.filter(producto=producto).aggregate(
                        total=Coalesce(Sum('metraje'), Decimal('0'))
                    )['total'] or Decimal('0')

                    creados = 0
                    duplicados = 0
                    excedidos = 0
                    codigos_duplicados = 0
                    metraje_excedido = 0

                    for item in unicos:
                        codigo_equipo = item['codigo']
                        valor = item['valor']

                        if es_cable_metraje:
                            metraje_item = _parse_metraje(valor)
                            if not codigo_equipo:
                                duplicados += 1
                                continue
                            if metraje_item <= 0:
                                duplicados += 1
                                continue

                            if codigo_equipo and DetalleIdentificador.objects.filter(codigo_individual__iexact=codigo_equipo).exists():
                                codigos_duplicados += 1
                                continue

                            if metraje_registrado + metraje_item > (producto.metraje or Decimal('0')):
                                metraje_excedido += 1
                                continue

                            obj, creado = DetalleIdentificador.objects.get_or_create(
                                producto=producto,
                                tipo='CABLE',
                                valor=codigo_equipo,
                                defaults={
                                    'codigo_individual': codigo_equipo,
                                    'metraje': metraje_item,
                                },
                            )
                            if obj and creado:
                                creados += 1
                                metraje_registrado += metraje_item
                            else:
                                duplicados += 1
                            continue

                        if es_codigo_unico:
                            codigo_base = codigo_equipo or valor
                            codigo_base = _normalizar_identificador(codigo_base)
                            if not codigo_base:
                                duplicados += 1
                                continue

                            if capacidad_disponible <= 0:
                                excedidos += 1
                                continue

                            if DetalleIdentificador.objects.filter(codigo_individual__iexact=codigo_base).exists():
                                codigos_duplicados += 1
                                continue

                            obj, creado = DetalleIdentificador.objects.get_or_create(
                                producto=producto,
                                tipo='SERIAL',
                                valor=codigo_base,
                                defaults={'codigo_individual': codigo_base},
                            )
                            if obj and creado:
                                creados += 1
                                capacidad_disponible -= 1
                            else:
                                duplicados += 1
                            continue

                        if not valor:
                            duplicados += 1
                            continue

                        if capacidad_disponible <= 0:
                            excedidos += 1
                            continue

                        if codigo_equipo and DetalleIdentificador.objects.filter(codigo_individual__iexact=codigo_equipo).exists():
                            codigos_duplicados += 1
                            continue

                        obj, creado = DetalleIdentificador.objects.get_or_create(
                            producto=producto,
                            tipo=tipo_masivo,
                            valor=valor,
                            defaults={'codigo_individual': codigo_equipo or None},
                        )
                        if obj and creado:
                            creados += 1
                            capacidad_disponible -= 1
                        else:
                            if codigo_equipo and not obj.codigo_individual:
                                obj.codigo_individual = codigo_equipo
                                obj.save(update_fields=['codigo_individual'])
                            duplicados += 1

                    if es_cable_metraje:
                        messages.success(
                            request,
                            f'Importacion finalizada. Creados: {creados}, Duplicados: {duplicados}, Metraje excedido: {metraje_excedido}, Codigos duplicados: {codigos_duplicados}.'
                        )
                    else:
                        messages.success(
                            request,
                            f'Importacion finalizada. Creados: {creados}, Duplicados: {duplicados}, Excedidos por stock: {excedidos}, Codigos duplicados: {codigos_duplicados}.'
                        )
                    return redirect(detalle_url)

        else:
            if es_cable_metraje:
                codigo_equipo = _normalizar_identificador(request.POST.get('codigo_individual'))
                metraje_item = _parse_metraje(request.POST.get('metraje_individual'))

                if not codigo_equipo:
                    messages.error(request, 'Debe ingresar un codigo unico para el cable/fibra.')
                    return redirect(detalle_url)

                if metraje_item <= 0:
                    messages.error(request, 'Debe ingresar un metraje mayor a cero por unidad.')
                    return redirect(detalle_url)

                if DetalleIdentificador.objects.filter(codigo_individual__iexact=codigo_equipo).exists():
                    messages.error(request, 'El codigo unico ya existe. Debe ingresar uno diferente.')
                    return redirect(detalle_url)

                metraje_registrado = DetalleIdentificador.objects.filter(producto=producto).aggregate(
                    total=Coalesce(Sum('metraje'), Decimal('0'))
                )['total'] or Decimal('0')

                if metraje_registrado + metraje_item > (producto.metraje or Decimal('0')):
                    restante = (producto.metraje or Decimal('0')) - metraje_registrado
                    messages.error(request, f'No hay metraje suficiente. Disponible para registrar: {restante} m.')
                    return redirect(detalle_url)

                try:
                    DetalleIdentificador.objects.create(
                        producto=producto,
                        tipo='CABLE',
                        codigo_individual=codigo_equipo,
                        valor=codigo_equipo,
                        metraje=metraje_item,
                    )
                    messages.success(request, 'Cable/fibra registrado correctamente con codigo unico y metraje.')
                    return redirect(detalle_url)
                except IntegrityError:
                    messages.error(request, 'Ese codigo ya fue registrado para este producto.')

            elif es_codigo_unico:
                codigo_equipo = _normalizar_identificador(request.POST.get('codigo_individual'))

                if not codigo_equipo:
                    messages.error(request, 'Debe ingresar un codigo unico.')
                    return redirect(detalle_url)

                if DetalleIdentificador.objects.filter(codigo_individual__iexact=codigo_equipo).exists():
                    messages.error(request, 'El codigo unico ya existe. Debe ingresar uno diferente.')
                    return redirect(detalle_url)

                cantidad_registrada = DetalleIdentificador.objects.filter(producto=producto).count()
                if cantidad_registrada >= producto.stock:
                    messages.error(request, 'No puede registrar mas codigos que el stock disponible.')
                    return redirect(detalle_url)

                try:
                    DetalleIdentificador.objects.create(
                        producto=producto,
                        tipo='SERIAL',
                        codigo_individual=codigo_equipo,
                        valor=codigo_equipo,
                        metraje=0,
                    )
                    messages.success(request, 'Codigo unico registrado correctamente.')
                    return redirect(detalle_url)
                except IntegrityError:
                    messages.error(request, 'Ese codigo ya fue registrado para este producto.')

            else:
                form_identificador = DetalleIdentificadorForm(request.POST)
                if form_identificador.is_valid():
                    detalle = form_identificador.save(commit=False)
                    detalle.producto = producto

                    if detalle.tipo not in ['SERIAL', 'MAC']:
                        messages.error(request, 'Tipo de identificador invalido para equipo activo.')
                        return redirect(detalle_url)

                    if not (detalle.valor or '').strip():
                        messages.error(request, 'Debe ingresar un Serial o MAC.')
                        return redirect(detalle_url)

                    codigo_equipo = (detalle.codigo_individual or '').strip()
                    if codigo_equipo and DetalleIdentificador.objects.filter(codigo_individual__iexact=codigo_equipo).exists():
                        messages.error(request, 'El codigo unico ya existe. Debe ingresar uno diferente.')
                        return redirect(detalle_url)

                    cantidad_registrada = DetalleIdentificador.objects.filter(producto=producto).count()
                    if cantidad_registrada >= producto.stock:
                        messages.error(request, 'No puede registrar mas identificadores que el stock disponible.')
                    else:
                        try:
                            detalle.save()
                            messages.success(request, 'Identificador registrado correctamente.')
                            return redirect(detalle_url)
                        except IntegrityError:
                            messages.error(request, 'Ese Serial/MAC ya fue registrado para este producto.')

    form_identificador = DetalleIdentificadorForm()
    if es_activo:
        form_identificador.fields['tipo'].choices = [
            ('SERIAL', 'Serial'),
            ('MAC', 'MAC'),
        ]

    # Reparacion conservadora para casos heredados:
    # si solo una bobina/cable no asignada quedo en 0 m pero existe metraje
    # residual en el producto, se restaura ese residual a dicha bobina.
    if es_cable_metraje:
        cables_qs = DetalleIdentificador.objects.filter(producto=producto, tipo='CABLE')
        suma_cables = cables_qs.aggregate(total=Coalesce(Sum('metraje'), Decimal('0')))['total'] or Decimal('0')
        residual = (producto.metraje or Decimal('0')) - suma_cables

        if residual > Decimal('0.01'):
            candidatos_reparar = list(
                cables_qs.filter(metraje__lte=0, asignacion_detalle__isnull=True).order_by('-fecha_registro')
            )
            if len(candidatos_reparar) == 1:
                candidato = candidatos_reparar[0]
                candidato.metraje = residual
                candidato.save(update_fields=['metraje'])

    identificadores_qs = DetalleIdentificador.objects.filter(producto=producto).order_by('-fecha_registro')
    cantidad_registrada = identificadores_qs.count()

    identificadores_asignados = set(
        AsignacionIdentificador.objects.filter(
            identificador__producto=producto
        ).values_list('identificador_id', flat=True)
    )
    identificadores_baja = set(
        BajaIdentificador.objects.filter(
            identificador__producto=producto
        ).values_list('identificador_id', flat=True)
    )

    identificadores = []
    for item in identificadores_qs:
        es_cable = item.tipo == 'CABLE'
        estado = 'Bodega'
        if item.id in identificadores_asignados:
            estado = 'Cuadrilla'
        elif item.id in identificadores_baja:
            if es_cable and (item.metraje or Decimal('0')) > 0:
                estado = 'Bodega'
            else:
                estado = 'Utilizado'

        if filtro_estado and estado.lower() != filtro_estado:
            continue

        identificadores.append({
            'tipo': item.get_tipo_display(),
            'codigo_individual': item.codigo_individual,
            'valor': item.valor,
            'metraje': item.metraje,
            'fecha_registro': item.fecha_registro,
            'estado': estado,
        })

    cantidad_mostrada = len(identificadores)

    es_construccion = producto.bodega == Producto.BODEGA_CONSTRUCCION

    return render(request, 'inventario/detalle_producto.html', {
        'producto': producto,
        'es_activo': es_activo,
        'es_fibra': es_fibra,
        'es_metraje': es_metraje,
        'es_cable_metraje': es_cable_metraje,
        'es_codigo_unico': es_codigo_unico,
        'es_construccion': es_construccion,
        'form_identificador': form_identificador,
        'identificadores': identificadores,
        'cantidad_registrada': cantidad_registrada,
        'cantidad_mostrada': cantidad_mostrada,
        'filtro_estado': filtro_estado,
    })

@role_required('admin', 'bodega')
def dashboard(request):

    total_productos = Producto.objects.count()
    total_cuadrillas = Cuadrilla.objects.count()

    total_stock = Producto.objects.aggregate(
        total=Sum('stock')
    )['total'] or 0

    total_asignado = AsignacionMaterial.objects.aggregate(
        total=Sum('cantidad')
    )['total'] or 0

    porcentaje_asignado = 0
    if total_stock > 0:
        porcentaje_asignado = round((total_asignado / total_stock) * 100, 1)
    porcentaje_asignado_barra = min(porcentaje_asignado, 100)

    productos_stock_bajo = Producto.objects.filter(stock__lte=F('stock_minimo')).order_by('nombre')

    return render(request, "dashboard.html", {

        "total_productos": total_productos,
        "total_cuadrillas": total_cuadrillas,
        "total_stock": total_stock,
        "total_asignado": total_asignado,
        "porcentaje_asignado": porcentaje_asignado,
        "porcentaje_asignado_barra": porcentaje_asignado_barra,
        "productos_stock_bajo": productos_stock_bajo,
        "total_stock_bajo": productos_stock_bajo.count(),
    })


@role_required('admin', 'bodega')
def recursos_sistema(request):
    return render(request, 'inventario/recursos_sistema.html', {
        'snapshot': _resource_snapshot(),
    })


@role_required('admin', 'bodega')
def recursos_sistema_data(request):
    return JsonResponse(_resource_snapshot())

@role_required('admin', 'bodega')
def modelos_resumen(request):

    resumen = Producto.objects.values('modelo__nombre').annotate(
        total=Sum('stock')
    )

    return render(request, 'inventario/resumen_modelos.html', {
        'resumen': resumen
    })

@role_required('admin', 'bodega')
def inventario_modelos(request):

    productos = Producto.objects.select_related(
        'marca', 'modelo', 'categoria'
    ).filter(
        stock__gt=0
    ).order_by(
        'marca__nombre', 'modelo__nombre', 'nombre'
    )

    modelos = []
    for producto in productos:
        categoria_nombre = producto.categoria.nombre if producto.categoria else ''
        activo = es_categoria_activa(categoria_nombre)

        marca_nombre = producto.marca.nombre if producto.marca else 'Sin marca'
        modelo_nombre = producto.modelo.nombre if producto.modelo else producto.nombre

        if activo:
            detalle_valor = producto.mac or producto.serial or 'Sin MAC/Serial cargado'
            tipo_detalle = 'Equipo activo'
        else:
            detalle_valor = producto.nombre
            tipo_detalle = 'Material pasivo'

        modelos.append({
            'id': producto.id,
            'nombre_producto': producto.nombre,
            'marca': marca_nombre,
            'modelo': modelo_nombre,
            'stock': producto.stock,
            'tipo_detalle': tipo_detalle,
            'detalle_valor': detalle_valor,
            'activo': activo,
        })

    return render(request, 'inventario/inventario_modelos.html', {
        'modelos': modelos
    })

@role_required('admin', 'bodega', 'cuadrilla')
def buscar_producto(request):

    codigo = (request.GET.get("codigo") or '').strip()
    nombre = (request.GET.get("nombre") or '').strip()
    identificador = (request.GET.get("identificador") or '').strip()

    productos = None
    resultado_identificador = None
    resultados_identificadores = []

    detalles_qs = DetalleIdentificador.objects.none()

    if identificador:
        det = DetalleIdentificador.objects.filter(valor__iexact=identificador).select_related('producto').first()
        resultado_identificador = det
        detalles_qs = DetalleIdentificador.objects.filter(valor__icontains=identificador).select_related('producto').order_by('-fecha_registro')
        if det:
            productos = Producto.objects.filter(pk=det.producto.pk)
        else:
            productos = Producto.objects.none()
    elif codigo or nombre:
        productos = Producto.objects.all()
        if codigo:
            detalles_qs = DetalleIdentificador.objects.filter(
                codigo_individual__icontains=codigo
            ).select_related('producto').order_by('-fecha_registro')
            ids_por_codigo_individual = detalles_qs.values_list('producto_id', flat=True)
            productos = productos.filter(
                Q(codigo__icontains=codigo) | Q(pk__in=ids_por_codigo_individual)
            )
        if nombre:
            productos = productos.filter(nombre__icontains=nombre)

    detalle_ids = list(detalles_qs.values_list('id', flat=True))
    if detalle_ids:
        bajas_por_identificador = {}
        for baja_identificador in BajaIdentificador.objects.filter(
            identificador_id__in=detalle_ids
        ).select_related('baja__cliente', 'baja').order_by('identificador_id', '-baja__fecha'):
            bajas_por_identificador.setdefault(baja_identificador.identificador_id, baja_identificador)

        asignaciones_por_identificador = {}
        for asignacion_identificador in AsignacionIdentificador.objects.filter(
            identificador_id__in=detalle_ids
        ).select_related('asignacion__cuadrilla', 'asignacion').order_by('identificador_id', '-fecha'):
            asignaciones_por_identificador.setdefault(asignacion_identificador.identificador_id, asignacion_identificador)

        for detalle in detalles_qs:
            baja_identificador = bajas_por_identificador.get(detalle.id)
            asignacion_identificador = asignaciones_por_identificador.get(detalle.id)
            es_cable = detalle.tipo == 'CABLE'
            estado = 'Bodega'
            cliente = '-'
            tipo_uso = '-'
            fecha_movimiento = detalle.fecha_registro
            filtro_estado = 'bodega'

            if asignacion_identificador:
                estado = 'Cuadrilla'
                filtro_estado = 'cuadrilla'
                cliente = asignacion_identificador.asignacion.cuadrilla.nombre
                fecha_movimiento = asignacion_identificador.fecha
            elif baja_identificador:
                if es_cable and (detalle.metraje or Decimal('0')) > 0:
                    estado = 'Bodega'
                    filtro_estado = 'bodega'
                else:
                    estado = 'Utilizado'
                    filtro_estado = 'utilizado'
                    cliente = baja_identificador.baja.cliente_nombre or (
                        baja_identificador.baja.cliente.nombre if baja_identificador.baja.cliente else '-'
                    )
                    tipo_uso = baja_identificador.baja.tipo_uso or '-'
                    fecha_movimiento = baja_identificador.baja.fecha

            resultados_identificadores.append({
                'producto_id': detalle.producto_id,
                'producto_codigo': detalle.producto.codigo,
                'producto_nombre': detalle.producto.nombre,
                'tipo': detalle.get_tipo_display(),
                'codigo_individual': detalle.codigo_individual,
                'valor': detalle.valor,
                'estado': estado,
                'cliente': cliente,
                'tipo_uso': tipo_uso,
                'fecha_movimiento': fecha_movimiento,
                'filtro_estado': filtro_estado,
            })

    return render(request, "inventario/buscar_producto.html", {
        "productos": productos,
        "codigo": codigo,
        "nombre": nombre,
        "identificador": identificador,
        "resultado_identificador": resultado_identificador,
        "resultados_identificadores": resultados_identificadores,
    })

@role_required('admin', 'bodega')
def reporte_inventario_pdf(request):

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario.pdf"'
    pdf = canvas.Canvas(response, pagesize=letter)
    page_width, page_height = letter
    logo_path = FilePath(settings.BASE_DIR) / 'static' / 'img' / 'logo_fastnett.png'

    def dibujar_encabezado():
        pdf.setFillColor(colors.HexColor('#1f2933'))
        pdf.rect(0, page_height - 55, page_width, 55, fill=1, stroke=0)

        if logo_path.exists():
            try:
                pdf.drawImage(str(logo_path), 40, page_height - 48, width=88, height=36, mask='auto')
            except Exception:
                pass

        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 15)
        pdf.drawString(140, page_height - 35, "Reporte de Inventario")

        pdf.setFont("Helvetica", 9)
        fecha_reporte = timezone.localtime().strftime('%d/%m/%Y %H:%M')
        pdf.drawRightString(page_width - 40, page_height - 35, f"Generado: {fecha_reporte}")

        pdf.setFillColor(colors.black)

    def dibujar_cabecera_tabla(y):
        pdf.setFillColor(colors.HexColor('#e5e7eb'))
        pdf.rect(35, y - 4, page_width - 70, 18, fill=1, stroke=0)

        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(40, y, "Codigo")
        pdf.drawString(100, y, "Producto")
        pdf.drawString(255, y, "Categoria")
        pdf.drawString(355, y, "Stock")
        pdf.drawString(400, y, "Minimo")
        pdf.drawString(455, y, "Estado")

    dibujar_encabezado()
    y = page_height - 80
    dibujar_cabecera_tabla(y)
    y -= 18

    productos = Producto.objects.select_related('categoria').order_by('nombre')
    pdf.setFont("Helvetica", 9)

    for producto in productos:
        if y < 55:
            pdf.showPage()
            dibujar_encabezado()
            y = page_height - 80
            dibujar_cabecera_tabla(y)
            y -= 18
            pdf.setFont("Helvetica", 9)

        estado = 'Stock bajo' if producto.stock <= producto.stock_minimo else 'OK'
        nombre_categoria = producto.categoria.nombre if producto.categoria else 'Sin categoria'

        pdf.drawString(40, y, str(producto.codigo or '-'))
        pdf.drawString(100, y, str((producto.nombre or '-')[:28]))
        pdf.drawString(255, y, str((nombre_categoria or '-')[:18]))
        pdf.drawRightString(390, y, str(producto.stock))
        pdf.drawRightString(440, y, str(producto.stock_minimo))

        if estado == 'Stock bajo':
            pdf.setFillColor(colors.red)
        else:
            pdf.setFillColor(colors.darkgreen)
        pdf.drawString(455, y, estado)
        pdf.setFillColor(colors.black)

        y -= 15

    pdf.setFont("Helvetica-Oblique", 8)
    pdf.drawString(40, 30, "Fastnett - Reporte automatico")

    pdf.save()

    return response


# ── Proveedores ──────────────────────────────────────────────────────────────

@role_required('admin', 'bodega')
def lista_proveedores(request):
    proveedores = Proveedor.objects.all()
    return render(request, 'inventario/proveedores.html', {'proveedores': proveedores})


@role_required('admin', 'bodega')
def crear_proveedor(request):
    form = ProveedorForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Proveedor creado correctamente.')
        return redirect('lista_proveedores')
    return render(request, 'inventario/crear_proveedor.html', {'form': form})


@role_required('admin', 'bodega')
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, pk=id)
    form = ProveedorForm(request.POST or None, instance=proveedor)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Proveedor actualizado correctamente.')
        return redirect('lista_proveedores')
    return render(request, 'inventario/editar_proveedor.html', {'form': form, 'proveedor': proveedor})


@role_required('admin', 'bodega')
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, pk=id)
    if request.method == 'POST':
        proveedor.delete()
        messages.success(request, 'Proveedor eliminado.')
        return redirect('lista_proveedores')
    messages.info(request, 'Confirme la eliminacion desde la lista de proveedores.')
    return redirect('lista_proveedores')